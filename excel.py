from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl.styles import Font

# grab the active worksheet

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'),
                     bottom=Side(style="thin"))

wb = Workbook()
ws = wb.active
wa = wb.create_sheet("a")
# property cell.border should be used instead of cell.style.border
ws.cell(row=3, column=2).border = thin_border
column_width = 10
fontsize = 11
font = Font(name="Bahnschrift SemiBold",size=fontsize)#underline="single"
ws.cell(15,2,value="asdasasdasdasdasd").font = font
for cell in ws["B":"B"]:
    if len(str(cell.value))+5 > column_width:
        column_width = (len(str(cell.value)))*(fontsize/11)+5
ws.column_dimensions["B"].width = column_width
wb.save('border_test.xlsx')
