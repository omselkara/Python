from tkinter import *
from tkinter.ttk import *
from tkinter import messagebox,filedialog
import time,os
import numpy as np
from copy import deepcopy
import openpyxl as pyx
from openpyxl.styles import numbers
from datetime import datetime

class main:
    def __init__(self):
        self.pencere = Tk()
        self.pencere.title("KÜTÜPHANE")
        self.pencere.resizable(0,0)
        self.resx = 700
        self.resy = 650
        entry = Entry(self.pencere,width=27)
        self.canvas = Canvas(self.pencere,width=self.resx,height=self.resy,highlightthickness=0,bg="#357D84")
        self.canvas.pack()
        self.bilgiler = {}
        self.başlıklar = []
        self.entrybilgi = []
        self.eskiiş = ""
        self.eskigünlük = ""
        self.eskibaşlıklar = []
        row = 25
        for y in range(row+1):
            self.canvas.create_rectangle(0,y*((self.resy-1)/row),175,(y+1)*((self.resy-1)/row),fill="#597487")
            self.canvas.create_line(0,y*((self.resy-1)/row),500,y*((self.resy-1)/row))
            if y<row:
                entry = Entry(self.canvas,width=27)
                self.başlıklar.append(entry)
                self.canvas.create_window(5,y*((self.resy-1)/row)+((self.resy-1)/row)/2,window=entry,anchor=W)
                self.eskibaşlıklar.append("")
                entry = Entry(self.canvas,width=53)
                self.entrybilgi.append(entry)
                self.canvas.create_window(325/2+175,y*((self.resy-1)/row)+((self.resy-1)/row)/2,window=entry)
        for x in range(2):
            self.canvas.create_line(x*500,0,x*500,self.resy)
        self.canvas.create_line(175,0,175,self.resy)
        self.güzergahkombo = Combobox(self.canvas,state="readonly")
        self.eskigüzergah = ""
        self.canvas.create_window(600,25,window=self.güzergahkombo)
        buton = Button(self.canvas,text="KAYDET",command=self.kayıt_yap)
        self.canvas.create_window(600,100,window=buton)
        buton = Button(self.canvas,text="SEÇİLİ KAYDI SİL",command=self.kayıt_sil)
        self.canvas.create_window(600,250,window=buton)
        buton = Button(self.canvas,text="EXCELDEN AL",command=self.al)
        self.canvas.create_window(600,467,window=buton)
        buton = Button(self.canvas,text="TÜM KAYITLARI SİL",command=self.temizle)
        self.canvas.create_window(600,550,window=buton)
        buton = Button(self.canvas,text="EXCELE AKTAR",command=self.aktar)
        self.canvas.create_window(600,325,window=buton)
        buton = Button(self.canvas,text="YEDEK AL",command=self.yedek_al)
        self.canvas.create_window(600,375,window=buton)
        buton = Button(self.canvas,text="YEDEKTEN GERİ YÜKLE",command=self.yedekten_al)
        self.canvas.create_window(600,425,window=buton)
        self.load_data()
        
    def update(self):
        self.canvas.update()
        for i in range(len(self.entrybilgi)):
            if self.başlıklar[i].get()!=self.eskibaşlıklar[i]:
                self.eskibaşlıklar[i] = self.başlıklar[i].get()
                self.save_data()
                break
        self.güzergahkombo["values"] = [""]+[i for i in self.bilgiler]
        if self.güzergahkombo.get()!=self.eskigüzergah:
            self.eskigüzergah = self.güzergahkombo.get()
            for index,i in enumerate(self.entrybilgi):
                i.delete(0,"end")
                if self.güzergahkombo.get()!="":
                    i.insert(0,self.bilgiler[self.güzergahkombo.get()][index])

    def yedek_al(self):
        f = filedialog.asksaveasfile(mode='w', defaultextension=".npz")
        if f is None:
            return
        name = f.name
        f.close()
        
        self.save_data(name)
        
    def yedekten_al(self):
        filename = filedialog.askopenfilename(title = "Dosya Seç", filetypes = [("Yedek Dosyalar",".npz")])
        if filename=="":
            return
        
        self.load_data(filename)
    
    def load_data(self,name="data.npz"):
        try:
            file = np.load(name,allow_pickle=True)
        except:
            return
        başlık = file["başlık"].tolist()
        for i in range(len(başlık)):
            self.başlıklar[i].delete(0,"end")
            self.başlıklar[i].insert(0,başlık[i])
            self.eskibaşlıklar[i] = başlık[i]
        self.bilgiler = file["bilgiler"].tolist()
        
    def save_data(self,name="data.npz"):
        di = {}
        di["başlık"] = [i.get() for i in self.başlıklar]
        di["bilgiler"] = self.bilgiler
        file = np.savez(name,**di)
        
    def kayıt_yap(self):
        if self.entrybilgi[0].get()=="":
            messagebox.showerror("HATA",f"{self.başlıklar[0].get()} GİR")
            return
        self.bilgiler[self.entrybilgi[0].get()] = [i.get() for i in self.entrybilgi]
        self.save_data()
        self.güzergahkombo.set(self.entrybilgi[0].get())
        messagebox.showinfo("BİLGİ","KAYDEDİLDİ")

    def kayıt_sil(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA",f"{self.başlıklar[0].get()} SEÇ")
            return
        del self.bilgiler[self.güzergahkombo.get()]
        messagebox.showinfo("BİLGİ",f"{self.güzergahkombo.get()} SİLİNDİ")
        self.güzergahkombo.set("")
        for i in self.entrybilgi:
            i.delete(0,"end")

    def temizle(self):
        answer = messagebox.askquestion("EMİNMİSİN","EMİNMİSİN TÜM KAYITLAR SİLİNECEKTİR")
        if answer=="yes":
            self.bilgiler = {}
            self.save_data()
            messagebox.showinfo("BİLGİ",f"TÜM KAYITLAR SİLİNDİ")
            self.güzergahkombo.set("")
            for i in self.entrybilgi:
                i.delete(0,"end")

    def aktar(self):
        f = filedialog.asksaveasfile(mode='w', defaultextension=".xlsx")
        if f is None:
            return
        name = f.name
        f.close()
        
        wb = pyx.Workbook()

        ws = wb.active
        for x in range(len(self.başlıklar)):
            ws.cell(column=x+1,row=1,value=str(self.başlıklar[x].get()))
            for y,güz in enumerate(self.bilgiler):
                ws.cell(column=x+1,row=y+2,value=str(self.bilgiler[güz][x]))
        
        wb.save(name)
        wb.close()
        try:
            os.startfile(name)
        except:
            messagebox.showerror("HATA","EXCEL DOSYASINI KAPATMADINIZ")

    def al(self):
        filename = filedialog.askopenfilename(title = "Dosya Seç", filetypes = [("EXCEL DOSYALARI",".xlsx")])
        if filename=="":
            return
        wb = pyx.load_workbook(filename)
        ws = wb.active
        
        for x in range(len(self.başlıklar)):
            self.başlıklar[x].delete(0,"end")
            if ws.cell(column=x+1,row=1).value!=None:
                self.başlıklar[x].insert(0,str(ws.cell(column=x+1,row=1).value))
        y = 2
        while 1:
            if ws.cell(column=1,row=y).value==None:
                break
            self.bilgiler[str(ws.cell(column=1,row=y).value)] = [str(ws.cell(column=x+1,row=y).value) if ws.cell(column=x+1,row=y).value!=None else "" for x in range(len(self.başlıklar))]
            y += 1
        self.save_data()
        
            
pencere = main()
while 1:
    try:
        pencere.update()
    except:
        pass
    
    time.sleep(1/60)
