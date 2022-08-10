from tkinter import *
from tkinter.ttk import *
from tkinter import messagebox,filedialog
import time,os
import numpy as np
from copy import deepcopy
import openpyxl as pyx
from openpyxl.styles import numbers
from datetime import datetime

aylar = ["EYLÜL","EKİM","KASIM","ARALIK","OCAK","ŞUBAT","MART","NİSAN","MAYIS","HAZİRAN","TEMMUZ","AĞUSTOS"]
sözlük = {"başlıklar":["GÜZERGAH","YÜKLENİCİ","TC/VERGİ NO","GÜNLÜK FİYAT(KDVSİZ)","İŞ GÜNÜ","SÖZLEŞME TOP.FİY. (KDVSİZ)","VERGİ USULÜ","BAŞLANGIÇ-BİTİŞ","İHALE NUMARASI","SÖZLEŞME TARİHİ","TOP. ÖDENEN HAKEDİŞ(KDVSİZ)","KOMİSYON ONAY TARİHİ","TEKLİF TUTANAK KOM. BAŞK.","ÜYE","ÜYE","HATIRLATMA","HATIRLATMA","HATIRLATMA","HATIRLATMA","HATIRLATMA"]}
class bilgi_pencere:
    def __init__(self,parent):
        self.active = 0
        self.parent = parent
        self.pencere = Tk()
        self.pencere.title("BİLGİ")
        self.pencere.resizable(0,0)
        self.resx = 700
        self.resy = 566
        self.pencere.withdraw()
        self.canvas = Canvas(self.pencere,width=self.resx,height=self.resy,highlightthickness=0,bg="#BEBEBE")
        self.canvas.pack()
        self.bilgiler = []
        self.eskiiş = ""
        self.eskigünlük = ""
        for y in range(21):
            self.canvas.create_rectangle(0,y*((self.resy-1)/20),175,(y+1)*((self.resy-1)/20),fill="#9DC4D8")
            self.canvas.create_line(0,y*((self.resy-1)/20),500,y*((self.resy-1)/20))
            if y<20:
                entry = Entry(self.canvas,width=53)
                self.bilgiler.append(entry)
                self.canvas.create_window(325/2+175,y*((self.resy-1)/20)+((self.resy-1)/20)/2,window=entry)
                self.canvas.create_text(175/2,y*((self.resy-1)/20)+((self.resy-1)/20)/2,text=sözlük["başlıklar"][y])
        self.bilgiler[10].config(state="readonly")
        for x in range(2):
            self.canvas.create_line(x*500,0,x*500,self.resy)
        self.canvas.create_line(175,0,175,self.resy)
        self.güzergahkombo = Combobox(self.canvas,state="readonly")
        self.eskigüzergah = ""
        self.canvas.create_window(600,25,window=self.güzergahkombo)
        buton = Button(self.canvas,text="KAYDET",command=self.kayıt_yap)
        self.canvas.create_window(600,100,window=buton)
        buton = Button(self.canvas,text="SEÇİLİ KAYDI SİL",command=self.kayıt_sil)
        self.canvas.create_window(600,250,window=buton)
        buton = Button(self.canvas,text="TÜM KAYITLARI SİL",command=self.reset)
        self.canvas.create_window(600,500,window=buton)
        buton = Button(self.canvas,text="EXCELDEN AKTAR",command=self.browse)
        self.canvas.create_window(600,325,window=buton)
        buton = Button(self.canvas,text="YEDEK AL",command=self.yedek_al)
        self.canvas.create_window(600,375,window=buton)
        buton = Button(self.canvas,text="YEDEKTEN GERİ YÜKLE",command=self.yedek_yükle)
        self.canvas.create_window(600,425,window=buton)
        self.pencere.protocol("WM_DELETE_WINDOW", self.hide)
        
    def show(self):
        self.active = 1
        self.pencere.deiconify()
        self.parent.hide()
        
    def hide(self):
        self.active = 0
        self.pencere.withdraw()
        self.parent.show()
        self.güzergahkombo.set("")
        
    def update(self):
        self.pencere.update()
        self.güzergahkombo["values"] = [""]+[i[0] for i in self.parent.bilgiler.items()]
        if self.eskigünlük!=self.bilgiler[3].get():
            self.eskigünlük = self.bilgiler[3].get()
            if self.bilgiler[3].get()!="" and self.bilgiler[4].get()!="":
                try:
                    sonuç = eval(self.bilgiler[3].get())*float(self.bilgiler[4].get())
                    self.bilgiler[5].delete(0,"end")
                    self.bilgiler[5].insert(0,trunc(sonuç))
                except:
                    pass
        if self.eskiiş!=self.bilgiler[4].get():
            self.eskiiş = self.bilgiler[4].get()
            if self.bilgiler[3].get()!="" and self.bilgiler[4].get()!="":
                try:
                    sonuç = eval(self.bilgiler[3].get())*float(self.bilgiler[4].get())
                    self.bilgiler[5].delete(0,"end")
                    self.bilgiler[5].insert(0,trunc(sonuç))
                except:
                    pass
            
        if self.eskigüzergah!=self.güzergahkombo.get():
            self.eskigüzergah = self.güzergahkombo.get()
            for index,i in enumerate(self.bilgiler):
                i.delete(0,"end")
                if self.güzergahkombo.get()!="":
                    i.insert(0,self.parent.bilgiler[self.güzergahkombo.get()][index])
            self.bilgiler[10].config(state="enabled")
            self.bilgiler[10].delete(0,"end")
            if self.güzergahkombo.get()!="":
                top = 0
                for i in range(12):
                    top += self.parent.hesaplar[self.güzergahkombo.get()][i][0]*self.parent.hesaplar[self.güzergahkombo.get()][i][1]
                self.bilgiler[10].insert(0,trunc(top))
            self.bilgiler[10].config(state="readonly")
    def kayıt_yap(self):
        if self.bilgiler[0].get()=="":
            messagebox.showerror("HATA","GÜZERGAH BOŞ OLAMAZ")
            return
        liste = []
        for i in self.bilgiler:
            liste.append(i.get())
            i.delete(0,"end")
        if not liste[0] in self.parent.bilgiler:
            self.parent.hesaplar[liste[0]] = [[0.0 for i in range(5)] for i in range(12)]
        self.parent.bilgiler[liste[0]] = deepcopy(liste)
        self.parent.yıllar[liste[0]] = [["",""] for i in range(12)]
        self.güzergahkombo.set("")
        self.parent.save_data()
        
    def kayıt_sil(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA","KAYIT SEÇİLMEDİ")
            return
        answer = messagebox.askquestion("EMİNMİSİN","EMİNMİSİN SEÇİLİ KAYIT SİLİNECEKTİR")
        if answer=="yes":
            del self.parent.bilgiler[self.güzergahkombo.get()]
            del self.parent.yıllar[self.güzergahkombo.get()]
            del self.parent.hesaplar[self.güzergahkombo.get()]
            self.güzergahkombo.set("")
            self.parent.save_data()

    def browse(self):
        filename = filedialog.askopenfilename(title = "Dosya Seç", filetypes = [("Excel Files",".xlsx")])
        if filename=="":
            return
        wb = pyx.load_workbook(filename)
        ws = wb.active
        y = 1
        while 1:
            if ws.cell(column=1,row=y).value==None:
                break
            liste = ["" for i in range(len(sözlük["başlıklar"]))]
            print(str(ws.cell(column=12,row=y).value))
            for i in range(len(sözlük["başlıklar"])):
                if ws.cell(column=i+1,row=y).value!=None:
                    liste[i] = ws.cell(column=i+1,row=y).value
            if not liste[0] in self.parent.bilgiler:
                self.parent.hesaplar[liste[0]] = [[0.0 for i in range(5)] for i in range(12)]
            try:
                if liste[5]=="":
                    liste[5] = trunc(float(liste[3])*float(liste[4]))
            except:
                liste[5] = ""
            self.parent.bilgiler[liste[0]] = deepcopy(liste)
            self.parent.yıllar[liste[0]] = [["",""] for i in range(12)]
            y += 1
        self.parent.save_data()
        wb.close()

    def reset(self):
        answer = messagebox.askquestion("EMİNMİSİN","EMİNMİSİN HERŞEY SIFIRLANACAKTIR")
        if answer=="yes":
            self.parent.bilgiler = {}
            self.parent.ekbilgiler = {"müdür":"","ilçemüdür":"","teslim":"","daire":"","birim":"","olur":""}
            self.parent.yıllar = {}
            self.parent.hesaplar = {}
            self.parent.save_data()
            self.parent.pencereler["hakediş"].müdür.delete(0,"end")
            self.parent.pencereler["hakediş"].ilçemüdür.delete(0,"end")
            self.parent.pencereler["hakediş"].eskimüdür = ""
            self.parent.pencereler["hakediş"].eskiilçemüdür = ""
            self.parent.pencereler["icmal"].teslim.delete(0,"end")
            self.parent.pencereler["icmal"].daire.delete(0,"end")
            self.parent.pencereler["icmal"].birim.delete(0,"end")
            self.parent.pencereler["icmal"].olur.delete(0,"end")
            self.parent.pencereler["icmal"].eskiolur = ""
            self.parent.pencereler["icmal"].eskibirim = ""
            self.parent.pencereler["icmal"].eskidaire = ""
            self.parent.pencereler["icmal"].eskiteslim = ""

    def yedek_al(self):
        f = filedialog.asksaveasfile(mode='w', defaultextension=".npz")
        if f is None:
            return
        name = f.name
        f.close()
        
        file = {}
        file["bilgiler"] = self.parent.bilgiler
        file["hesaplar"] = self.parent.hesaplar
        file["müdürler"] = self.parent.ekbilgiler
        file["yıllar"] = self.parent.yıllar
        np.savez(name,**file)

    def yedek_yükle(self):
        filename = filedialog.askopenfilename(title = "Dosya Seç", filetypes = [("Yedek Dosyalar",".npz")])
        if filename=="":
            return
        file = np.load(filename,allow_pickle=True)
        self.parent.bilgiler = file["bilgiler"].tolist()
        self.parent.hesaplar = file["hesaplar"].tolist()
        self.parent.ekbilgiler = file["müdürler"].tolist()
        self.parent.yıllar = file["yıllar"].tolist()
        self.parent.pencereler["hakediş"].müdür.delete(0,"end")
        self.parent.pencereler["hakediş"].ilçemüdür.delete(0,"end")
        self.parent.pencereler["icmal"].teslim.delete(0,"end")
        self.parent.pencereler["icmal"].daire.delete(0,"end")
        self.parent.pencereler["icmal"].birim.delete(0,"end")
        self.parent.pencereler["icmal"].olur.delete(0,"end")
        self.parent.pencereler["hakediş"].müdür.insert(0,self.parent.ekbilgiler["müdür"])
        self.parent.pencereler["hakediş"].ilçemüdür.insert(0,self.parent.ekbilgiler["ilçemüdür"])
        self.parent.pencereler["icmal"].teslim.insert(0,self.parent.ekbilgiler["teslim"])
        self.parent.pencereler["icmal"].daire.insert(0,self.parent.ekbilgiler["daire"])
        self.parent.pencereler["icmal"].birim.insert(0,self.parent.ekbilgiler["birim"])
        self.parent.pencereler["icmal"].olur.insert(0,self.parent.ekbilgiler["olur"])
        self.parent.pencereler["hakediş"].eskimüdür = self.parent.ekbilgiler["müdür"]
        self.parent.pencereler["hakediş"].eskiilçemüdür = self.parent.ekbilgiler["ilçemüdür"]
        self.parent.pencereler["icmal"].eskiteslim = self.parent.ekbilgiler["teslim"]
        self.parent.pencereler["icmal"].eskidaire = self.parent.ekbilgiler["daire"]
        self.parent.pencereler["icmal"].eskibirim = self.parent.ekbilgiler["birim"]
        self.parent.pencereler["icmal"].eskiolur = self.parent.ekbilgiler["olur"]
        self.parent.save_data()
        
class hakediş:
    def __init__(self,parent):
        self.parent = parent
        self.active = 0
        self.resx = 700
        self.resy = 550
        self.pencere = Tk()
        self.pencere.resizable(0,0)
        self.pencere.withdraw()
        self.pencere.title("HAKEDİŞ")
        self.pencere.protocol("WM_DELETE_WINDOW",self.hide)
        self.canvas = Canvas(self.pencere,width=self.resx,height=self.resy,highlightthickness=0,bg="#BEBEBE")
        self.canvas.pack()
        self.güzergahkombo = Combobox(self.canvas,state="readonly",width=45)
        self.eskigüzergah = ""
        self.canvas.create_window(100,10,window=self.güzergahkombo,anchor=NW)
        self.aykombo = Combobox(self.canvas,state="readonly",width=10)
        self.eskiay = ""
        self.canvas.create_window(470,10,window=self.aykombo,anchor=NW)
        self.aykombo["values"] = aylar
        self.canvas.create_text(5,10,text="GÜZERGAH SEÇ :",anchor=NW)
        self.canvas.create_text(420,10,text="AY SEÇ :",anchor=NW)
        self.usüllabel = self.canvas.create_text(250,50,text="VERGİ USULÜ :",anchor=NW)
        self.canvas.create_text(400,50,text="HAKEDİŞ TARİHİ :",anchor=NW)
        self.canvas.create_text(570,50,text="HAKEDİŞ NO :",anchor=NW)
        self.yüklenicilabel = self.canvas.create_text(5,50,text="YÜKLENİCİ :",anchor=NW)
        self.günlüklabel = self.canvas.create_text(5,110,text="GÜNLÜK FİYAT(KDVSİZ) :",anchor=NW)
        self.canvas.create_text(260,110,text="KDV :",anchor=NW)
        self.canvas.create_text(370,110,text="TEVKİFAT :",anchor=NW)
        self.canvas.create_text(510,110,text="DAMGA VERGİSİ :",anchor=NW)
        self.canvas.create_text(15,170,text="GÜN GİR :",anchor=NW)
        self.canvas.create_text(160,170,text="FİYAT GİR :",anchor=NW)
        self.canvas.create_text(self.resx/2,300,text="AYLIK ÇİZELGE HAKEDİŞ")
        self.canvas.create_text(self.resx/2,390,text="AYLIK ÇİZELGE KESİNTİLER")
        self.hakediş = self.canvas.create_text(300,170,text="HAKEDİŞ :",anchor=NW,fill="blue")
        self.canvas.create_text(50,520,text="ŞUBE MÜDÜRÜ :",anchor=NW)
        self.müdür = Entry(self.canvas,width=25)
        self.canvas.create_window(145,518,window=self.müdür,anchor=NW)
        self.canvas.create_text(325,520,text="İLÇE MİLLİ EĞİTİM MÜDÜRÜ :",anchor=NW)
        self.ilçemüdür = Entry(self.canvas,width=25)
        self.canvas.create_window(485,518,window=self.ilçemüdür,anchor=NW)
        self.müdür.insert(0,self.parent.ekbilgiler["müdür"])
        self.ilçemüdür.insert(0,self.parent.ekbilgiler["ilçemüdür"])
        self.eskimüdür = self.parent.ekbilgiler["müdür"]
        self.eskiilçemüdür = self.parent.ekbilgiler["ilçemüdür"]
        self.oranlar = []
        self.hesaplar = []
        e = Entry(self.canvas,width=8)
        self.oranlar.append(e)
        self.canvas.create_window(300,108,window=e,anchor=NW)
        e = Entry(self.canvas,width=8)
        self.oranlar.append(e)
        self.canvas.create_window(440,108,window=e,anchor=NW)
        e = Entry(self.canvas,width=8)
        self.oranlar.append(e)
        self.canvas.create_window(615,108,window=e,anchor=NW)
        self.tarih = Entry(self.canvas,width=9)
        self.canvas.create_window(500,50,window=self.tarih,anchor=NW)
        self.no = Entry(self.canvas,width=4)
        self.canvas.create_window(650,50,window=self.no,anchor=NW)
        self.gün = Entry(self.canvas,width=8)
        self.canvas.create_window(85,170,window=self.gün,anchor=NW)
        self.fiyat = Entry(self.canvas,width=8)
        self.canvas.create_window(225,170,window=self.fiyat,anchor=NW)
        buton = Button(self.canvas,text="HESAPLA",command=self.hesapla)
        self.canvas.create_window(430,177,window=buton,anchor=W)
        buton = Button(self.canvas,text="KAYDET",command=self.kaydet)
        self.canvas.create_window(575,177,window=buton,anchor=W)
        buton = Button(self.canvas,text="HAKEDİŞ YAZDIR",command=self.yazdır)
        self.canvas.create_window(580,20,window=buton,anchor=W)
        buton = Button(self.canvas,text="AYLIK ÇİZELGE SİL",command=self.reset)
        self.canvas.create_window(self.resx/2,470,window=buton)
        self.values = [0 for i in range(7)]
        for y in range(3):
            self.canvas.create_line(0,y*15+250,self.resx,y*15+250)
            self.canvas.create_line(0,y*15+330,self.resx,y*15+330)
            self.canvas.create_line(0,y*15+410,self.resx,y*15+410)
        for x in range(7):
            self.canvas.create_line(x*((self.resx-1)/6),250,x*((self.resx-1)/6),280)
            if x<6:
                self.hesaplar.append(self.canvas.create_text(x*((self.resx-1)/6)+((self.resx-1)/6)/2,272.5,text="0.00"))
        self.canvas.itemconfigure(self.hesaplar[1],fill="blue")
        self.canvas.itemconfigure(self.hesaplar[5],fill="blue")
        self.aylıkhak = []
        self.aylıkkes = []
        for x in range(13):
            self.canvas.create_line(x*((self.resx-1)/12),410,x*((self.resx-1)/12),440)
            self.canvas.create_line(x*((self.resx-1)/12),330,x*((self.resx-1)/12),360)
            if x<12:
                self.canvas.create_text(x*((self.resx-1)/12)+((self.resx-1)/12)/2,337.5,text=aylar[x])
                self.aylıkhak.append(self.canvas.create_text(x*((self.resx-1)/12)+((self.resx-1)/12)/2,352.5,text="0.00"))
                self.canvas.create_text(x*((self.resx-1)/12)+((self.resx-1)/12)/2,417.5,text=aylar[x])
                self.aylıkkes.append(self.canvas.create_text(x*((self.resx-1)/12)+((self.resx-1)/12)/2,432.5,text="0.00"))
        self.hakediştop = self.canvas.create_text(10*((self.resx-1)/12),370,text="TOPLAM : 0.00")
        self.kesintitop = self.canvas.create_text(10*((self.resx-1)/12),450,text="TOPLAM : 0.00")
        self.canvas.create_text(0*((self.resx-1)/6)+((self.resx-1)/6)/2,257.5,text="KDV")
        self.canvas.create_text(1*((self.resx-1)/6)+((self.resx-1)/6)/2,257.5,text="TAHAKKUK")
        self.canvas.create_text(2*((self.resx-1)/6)+((self.resx-1)/6)/2,257.5,text="TEVKİFAT")
        self.canvas.create_text(3*((self.resx-1)/6)+((self.resx-1)/6)/2,257.5,text="DAMGA V.")
        self.canvas.create_text(4*((self.resx-1)/6)+((self.resx-1)/6)/2,257.5,text="KESİNTİ TOP.")
        self.canvas.create_text(5*((self.resx-1)/6)+((self.resx-1)/6)/2,257.5,text="YÜK. ÖDENECEK")

    def reset(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA","GÜZERGAH SEÇİLİ DEĞİL")
            return
        answer = messagebox.askquestion("EMİNMİSİN","SEÇİLİ KİŞİYE AİT TÜM AYLAR SİLİNECEKTİR")
        if answer=="yes":
            self.parent.hesaplar[self.güzergahkombo.get()] = [[0.0 for i in range(5)]for i in range(12)]
            self.parent.yıllar[self.güzergahkombo.get()] = [["",""]for i in range(12)]
            messagebox.showinfo("BAŞARILI","BAŞARIYLA SİLİNDİ")
        self.parent.save_data()
    def update(self):
        self.pencere.update()
        self.güzergahkombo["values"] = [i[0] for i in self.parent.bilgiler.items()]
        if self.eskimüdür!=self.müdür.get():
            self.eskimüdür = self.müdür.get()
            self.parent.ekbilgiler["müdür"] = self.müdür.get()
            self.parent.ekbilgiler["ilçemüdür"] = self.ilçemüdür.get()
            self.parent.save_data()
        if self.eskiilçemüdür!=self.ilçemüdür.get():
            self.eskiilçemüdür = self.ilçemüdür.get()
            self.parent.ekbilgiler["müdür"] = self.müdür.get()
            self.parent.ekbilgiler["ilçemüdür"] = self.ilçemüdür.get()
            self.parent.save_data()
        if self.güzergahkombo.get()!=self.eskigüzergah:
            self.eskigüzergah = self.güzergahkombo.get()
            self.gün.delete(0,"end")
            self.fiyat.delete(0,"end")
            self.values = [0.0 for i in range(7)]
            if self.güzergahkombo.get()!="":
                self.canvas.itemconfigure(self.yüklenicilabel,text="YÜKLENİCİ : "+str(self.parent.bilgiler[self.güzergahkombo.get()][1]))
                self.canvas.itemconfigure(self.günlüklabel,text="GÜNLÜK FİYAT(KDVSİZ) : "+str(self.parent.bilgiler[self.güzergahkombo.get()][3]))
                self.canvas.itemconfigure(self.usüllabel,text="VERGİ USULÜ : "+str(self.parent.bilgiler[self.güzergahkombo.get()][6]))
                if self.aykombo.get()!="":
                    self.tarih.delete(0,"end")
                    self.no.delete(0,"end")
                    self.tarih.insert(0,self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
                    self.no.insert(0,self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1])
                    if self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]!=0:
                        self.gün.insert(0,str(self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]))
                    if self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]!=0:
                        self.fiyat.insert(0,str(self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]))
                    if (self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]!=0 or self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]!=0):
                        for i in range(3):
                            self.oranlar[i].delete(0,"end")
                            if self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][i+2]!=0:
                                self.oranlar[i].insert(0,str(self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][i+2]))
                        self.hesapla()
                    else:
                        for a in range(12):
                            if self.parent.hesaplar[self.güzergahkombo.get()][a][0]!=0 or self.parent.hesaplar[self.güzergahkombo.get()][a][1]!=0:
                                for i in range(3):
                                    self.oranlar[i].delete(0,"end")
                                    if self.parent.hesaplar[self.güzergahkombo.get()][a][i+2]!=0:
                                        self.oranlar[i].insert(0,str(self.parent.hesaplar[self.güzergahkombo.get()][a][i+2]))
                                break
                else:
                    for a in range(12):
                        if self.parent.hesaplar[self.güzergahkombo.get()][a][0]!=0 or self.parent.hesaplar[self.güzergahkombo.get()][a][1]!=0:
                            for i in range(3):
                                self.oranlar[i].delete(0,"end")
                                if self.parent.hesaplar[self.güzergahkombo.get()][a][i+2]!=0:
                                    self.oranlar[i].insert(0,str(self.parent.hesaplar[self.güzergahkombo.get()][a][i+2]))
                            break
            else:
                self.canvas.itemconfigure(self.yüklenicilabel,text="YÜKLENİCİ :")
                self.canvas.itemconfigure(self.günlüklabel,text="GÜNLÜK FİYAT(KDVSİZ) :")
                self.canvas.itemconfigure(self.usüllabel,text="VERGİ USULÜ :")
        if self.aykombo.get()!=self.eskiay:
            self.eskiay = self.aykombo.get()
            self.gün.delete(0,"end")
            self.fiyat.delete(0,"end")
            self.values = [0.0 for i in range(7)]
            if self.güzergahkombo.get()!="":
                self.canvas.itemconfigure(self.yüklenicilabel,text="YÜKLENİCİ : "+str(self.parent.bilgiler[self.güzergahkombo.get()][1]))
                self.canvas.itemconfigure(self.günlüklabel,text="GÜNLÜK FİYAT(KDVSİZ) : "+str(self.parent.bilgiler[self.güzergahkombo.get()][3]))
                self.canvas.itemconfigure(self.usüllabel,text="VERGİ USULÜ : "+str(self.parent.bilgiler[self.güzergahkombo.get()][6]))
                if self.güzergahkombo.get()!="":
                    self.tarih.delete(0,"end")
                    self.no.delete(0,"end")
                    self.tarih.insert(0,self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
                    self.no.insert(0,self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1])
                    if self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]!=0:
                        self.gün.insert(0,str(self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]))
                    if self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]!=0:
                        self.fiyat.insert(0,str(self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]))
                    if self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]!=0 or self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]!=0:
                        for i in range(3):
                            self.oranlar[i].delete(0,"end")
                            if self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][i+2]!=0:
                                self.oranlar[i].insert(0,str(self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][i+2]))
                        self.hesapla()
            else:
                self.canvas.itemconfigure(self.yüklenicilabel,text="YÜKLENİCİ :")
                self.canvas.itemconfigure(self.günlüklabel,text="GÜNLÜK FİYAT(KDVSİZ) :")
                self.canvas.itemconfigure(self.usüllabel,text="VERGİ USULÜ :")
        if self.güzergahkombo.get()!="":
            haktop = 0
            kestop = 0
            for i in range(12):
                hak = round(self.parent.hesaplar[self.güzergahkombo.get()][i][0]*self.parent.hesaplar[self.güzergahkombo.get()][i][1],2)
                self.canvas.itemconfigure(self.aylıkhak[i],text=trunc(hak))
                haktop += hak
                sonuç = round(hak*self.parent.hesaplar[self.güzergahkombo.get()][i][3],2)
                sonuç += round(hak*self.parent.hesaplar[self.güzergahkombo.get()][i][4],2)
                self.canvas.itemconfigure(self.aylıkkes[i],text=trunc(sonuç))
                kestop += sonuç
            self.canvas.itemconfigure(self.hakediştop,text="TOPLAM : "+trunc(haktop))
            self.canvas.itemconfigure(self.kesintitop,text="TOPLAM : "+trunc(kestop))
        self.canvas.itemconfigure(self.hakediş,text="HAKEDİŞ : "+trunc(self.values[0]))
        for x in range(6):
            self.canvas.itemconfigure(self.hesaplar[x],text=trunc(self.values[x+1]))

    def kaydet(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA","GÜZERGAH SEÇ")
            return
        if self.aykombo.get()=="":
            messagebox.showerror("HATA","AY SEÇ")
            return
        if self.gün.get()=="":
            messagebox.showerror("HATA","TAŞINAN GÜN GİR")
            return
        if self.fiyat.get()=="":
            messagebox.showerror("HATA","GÜNLÜK FİYAT GİR")
            return
        if self.tarih.get()=="" or self.tarih.get().count(".")!=2:
            messagebox.showerror("HATA","TARİH GİR")
            return
        if self.no.get()=="":
            messagebox.showerror("HATA","HAKEDİŞ NO GİR")
            return
        values = [0,0,0,0,0,0,0]
        safkdv = safe(self.oranlar[0].get())
        if safkdv==-1:
            messagebox.showerror("HATA","KDV HATALI")
            return
        saftev = safe(self.oranlar[1].get())
        if saftev==-1:
            messagebox.showerror("HATA","TEVKİFAT HATALI")
            return
        safdam = safe(self.oranlar[2].get())
        if safdam==-1:
            messagebox.showerror("HATA","DAMGA VERGİSİ HATALI")
            return
        kdv = 0
        tev = 0
        dam = 0
        if safkdv!="":
            kdv = float(safkdv)
        if saftev!="":
            tev = float(saftev)
        if safdam!="":
            dam = float(safdam)
        values[0] = round(float(self.gün.get())*float(self.fiyat.get()),2)
        values[1] = round(kdv*values[0],2)
        values[2] = round(values[0]+values[1],2)
        values[3] = round(tev*values[0],2)
        values[4] = round(dam*values[0],2)
        values[5] = round(values[3]+values[4],2)
        values[6] = round(values[2]-values[5],2)
        for i in range(7):
            if self.values[i]!=values[i]:
                messagebox.showerror("HATA","HESAPLAMA YAPILMADI")
                return
        self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())] = [float(self.gün.get()),float(self.fiyat.get()),kdv,tev,dam]
        self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0] = self.tarih.get()
        self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1] = self.no.get()
        self.parent.save_data()
        
    def show(self):
        self.active = 1
        self.pencere.deiconify()
        self.parent.hide()
        
    def hide(self):
        self.active = 0
        self.pencere.withdraw()
        self.parent.show()
        self.güzergahkombo.set("")
        self.aykombo.set("")
        
    def hesapla(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA","GÜZERGAH SEÇ")
            return
        if self.gün.get()=="":
            messagebox.showerror("HATA","TAŞINAN GÜN GİR")
            return
        if self.fiyat.get()=="":
            messagebox.showerror("HATA","GÜNLÜK FİYAT GİR")
            return
        safkdv = safe(self.oranlar[0].get())
        if safkdv==-1:
            messagebox.showerror("HATA","KDV HATALI")
            return
        saftev = safe(self.oranlar[1].get())
        if saftev==-1:
            messagebox.showerror("HATA","TEVKİFAT HATALI")
            return
        safdam = safe(self.oranlar[2].get())
        if safdam==-1:
            messagebox.showerror("HATA","DAMGA VERGİSİ HATALI")
            return
        kdv = 0
        tev = 0
        dam = 0
        if safkdv!="":
            kdv = float(safkdv)
        if saftev!="":
            tev = float(saftev)
        if safdam!="":
            dam = float(safdam)
        self.values[0] = round(float(self.gün.get())*float(self.fiyat.get()),2)
        self.values[1] = round(kdv*self.values[0],2)
        self.values[2] = round(self.values[0]+self.values[1],2)
        self.values[3] = round(tev*self.values[0],2)
        self.values[4] = round(dam*self.values[0],2)
        self.values[5] = round(self.values[3]+self.values[4],2)
        self.values[6] = round(self.values[2]-self.values[5],2)
    def yazdır(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA","GÜZERGAH SEÇ")
            return
        if self.aykombo.get()=="":
            messagebox.showerror("HATA","AY SEÇ")
            return
        if self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]=="":
            messagebox.showerror("HATA",f"{self.aykombo.get()} AYINA KAYIT YAPILMADI")
            return
        wb = pyx.load_workbook("Form.xlsx")
        for i in reversed(range(len(wb.worksheets))):
            if wb.sheetnames[i]!="HAKEDİŞ":
                wb.remove(wb.worksheets[i])
        ws = wb["HAKEDİŞ"]
        ws.cell(column=1,row=2,value=self.güzergahkombo.get())
        ws.cell(column=6,row=2,value=int(self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]))
        ws.cell(column=1,row=3,value=self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
        top = 0
        for i in range(aylar.index(self.aykombo.get())+1):
            top += self.parent.hesaplar[self.güzergahkombo.get()][i][0]*self.parent.hesaplar[self.güzergahkombo.get()][i][1]
        hesap = self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())]
        ws.cell(column=4,row=4,value=top)
        ws.cell(column=4,row=6,value=top)
        sonuç = hesap[0]*hesap[1]
        ws.cell(column=4,row=8,value=top-sonuç)
        ws.cell(column=4,row=9,value=sonuç)
        ws.cell(column=4,row=10,value=sonuç*hesap[2])
        ws.cell(column=4,row=11,value=sonuç*hesap[2]+sonuç)
        ws.cell(column=4,row=13,value=sonuç*hesap[4])
        ws.cell(column=4,row=14,value=sonuç*hesap[3])
        ws.cell(column=4,row=23,value=sonuç*hesap[3]+sonuç*hesap[4])
        ws.cell(column=4,row=24,value=(sonuç*hesap[2]+sonuç)-(sonuç*hesap[3]+sonuç*hesap[4]))
        ws.cell(column=2,row=27,value=self.parent.bilgiler[self.güzergahkombo.get()][1])
        ws.cell(column=5,row=28,value=self.müdür.get())
        ws.cell(column=5,row=31,value=self.ilçemüdür.get())
        try:
            wb.save("hakedişform.xlsx")
            os.startfile("hakedişform.xlsx")
        except:
            messagebox.showerror("HATA","EXCEL DOSYASINI KAPATMADINIZ")
        

class icmal:
    def __init__(self,parent):
        self.parent = parent
        self.active = 0
        self.resx = 440
        self.resy = 500
        self.pencere = Tk()
        self.pencere.resizable(0,0)
        self.pencere.withdraw()
        self.pencere.title("TUTANAKLAR")
        self.pencere.protocol("WM_DELETE_WINDOW",self.hide)
        self.canvas = Canvas(self.pencere,width=self.resx,height=self.resy,highlightthickness=0,bg="#BEBEBE")
        self.canvas.pack()
        self.canvas.create_text(10,15,text="GÜZERGAH SEÇ :",anchor=NW)
        self.canvas.create_text(240,85,text="AY SEÇ :",anchor=NW)
        self.yüklenici = self.canvas.create_text(10,45,text="YÜKLENİCİ :",anchor=NW)
        self.canvas.create_text(130,365,text="TESLİM EDEN :",anchor=NE)
        self.canvas.create_text(130,395,text="MUHASEBE DAİRESİ :",anchor=NE)
        self.canvas.create_text(130,425,text="MUHASEBE BİRİMİ :",anchor=NE)
        self.canvas.create_text(130,455,text="KOM. OLUR TARİHİ :",anchor=NE)
        self.güzergahkombo = Combobox(self.canvas,width=40,state="readonly")
        self.canvas.create_window(110,13,window=self.güzergahkombo,anchor=NW)
        self.aykombo = Combobox(self.canvas,width=10,state="readonly")
        self.canvas.create_window(290,83,window=self.aykombo,anchor=NW)
        self.aykombo["values"] = aylar
        self.teslim = Entry(self.canvas,width=40)
        self.canvas.create_window(140,364,window=self.teslim,anchor=NW)
        self.daire = Entry(self.canvas,width=40)
        self.canvas.create_window(140,394,window=self.daire,anchor=NW)
        self.birim = Entry(self.canvas,width=40)
        self.canvas.create_window(140,424,window=self.birim,anchor=NW)
        self.olur = Entry(self.canvas,width=40)
        self.canvas.create_window(140,454,window=self.olur,anchor=NW)
        self.eskiteslim = self.parent.ekbilgiler["teslim"]
        self.eskidaire = self.parent.ekbilgiler["daire"]
        self.eskibirim = self.parent.ekbilgiler["birim"]
        self.eskiolur = self.parent.ekbilgiler["olur"]
        self.teslim.insert(0,self.parent.ekbilgiler["teslim"])
        self.daire.insert(0,self.parent.ekbilgiler["daire"])
        self.birim.insert(0,self.parent.ekbilgiler["birim"])
        self.olur.insert(0,self.parent.ekbilgiler["olur"])
        self.eskigüzergah = ""
        buton = Button(self.canvas,text="İCMAL YAZDIR",width=25,command=self.icmal_yazdır)
        self.canvas.create_window(85,125,window=buton,anchor=NW)
        buton = Button(self.canvas,text="ÖZET YAZDIR",width=25,command=self.özet_yazdır)
        self.canvas.create_window(85,175,window=buton,anchor=NW)
        buton = Button(self.canvas,text="TEKLİF TUTANAK YAZDIR",width=25,command=self.teklif_yazdır)
        self.canvas.create_window(85,225,window=buton,anchor=NW)
        buton = Button(self.canvas,text="TESLİM TUTANAK YAZDIR",width=25,command=self.teslim_yazdır)
        self.canvas.create_window(85,275,window=buton,anchor=NW)
    def show(self):
        self.active = 1
        self.pencere.deiconify()
        self.parent.hide()
        
    def hide(self):
        self.active = 0
        self.pencere.withdraw()
        self.güzergahkombo.set("")
        self.aykombo.set("")
        self.parent.show()

    def update(self):
        self.pencere.update()
        self.güzergahkombo["values"] = [i[0] for i in self.parent.bilgiler.items()]
        if self.güzergahkombo.get()!=self.eskigüzergah:
            self.eskigüzergah = self.güzergahkombo.get()
            if self.güzergahkombo.get()!="":
                self.canvas.itemconfigure(self.yüklenici,text="YÜKLENİCİ : "+self.parent.bilgiler[self.güzergahkombo.get()][1])
        if self.eskiteslim!=self.teslim.get():
            self.eskiteslim = self.teslim.get()
            self.parent.ekbilgiler["teslim"] = self.teslim.get()
            self.parent.ekbilgiler["daire"] = self.daire.get()
            self.parent.ekbilgiler["birim"] = self.birim.get()
            self.parent.ekbilgiler["olur"] = self.olur.get()
            self.parent.save_data()
        if self.eskidaire!=self.daire.get():
            self.eskidaire = self.daire.get()
            self.parent.ekbilgiler["teslim"] = self.teslim.get()
            self.parent.ekbilgiler["daire"] = self.daire.get()
            self.parent.ekbilgiler["birim"] = self.birim.get()
            self.parent.ekbilgiler["olur"] = self.olur.get()
            self.parent.save_data()
        if self.eskibirim!=self.birim.get():
            self.eskibirim = self.birim.get()
            self.parent.ekbilgiler["teslim"] = self.teslim.get()
            self.parent.ekbilgiler["daire"] = self.daire.get()
            self.parent.ekbilgiler["birim"] = self.birim.get()
            self.parent.ekbilgiler["olur"] = self.olur.get()
            self.parent.save_data()
        if self.eskiolur!=self.olur.get():
            self.eskiolur = self.olur.get()
            self.parent.ekbilgiler["teslim"] = self.teslim.get()
            self.parent.ekbilgiler["daire"] = self.daire.get()
            self.parent.ekbilgiler["birim"] = self.birim.get()
            self.parent.ekbilgiler["olur"] = self.olur.get()
            self.parent.save_data()
        
    def icmal_yazdır(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA","GÜZERGAH SEÇ")
            return
        if self.aykombo.get()=="":
            messagebox.showerror("HATA","AY SEÇ")
            return
        if self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]=="":
            messagebox.showerror("HATA",f"{self.aykombo.get()} AYINA KAYIT YAPILMADI")
            return
        wb = pyx.load_workbook("Form.xlsx")
        for i in reversed(range(len(wb.worksheets))):
            if wb.sheetnames[i]!="İCMAL":
                wb.remove(wb.worksheets[i])
        ws = wb["İCMAL"]
        ws.cell(column=1,row=2,value=self.parent.bilgiler[self.güzergahkombo.get()][1]+" "+self.güzergahkombo.get())
        ws.cell(column=6,row=2,value=int(self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]))
        y = 4
        top = 0
        hakediş = (self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]*
                   self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1])
        son = 0
        eski = 0
        for i in range(aylar.index(self.aykombo.get())+1):
            if (self.parent.hesaplar[self.güzergahkombo.get()][i][0]!=0 or self.parent.hesaplar[self.güzergahkombo.get()][i][1]!=0):
                tahakkuk = ((self.parent.hesaplar[self.güzergahkombo.get()][i][0]*self.parent.hesaplar[self.güzergahkombo.get()][i][1])*
                            self.parent.hesaplar[self.güzergahkombo.get()][i][2]+
                            (self.parent.hesaplar[self.güzergahkombo.get()][i][0]*self.parent.hesaplar[self.güzergahkombo.get()][i][1]))
                top += tahakkuk
                yıl = self.parent.yıllar[self.güzergahkombo.get()][i][0]
                yıl = yıl[(len(yıl)-yıl[::-1].index(".")):]
                ws.cell(column=2,row=y,value=self.güzergahkombo.get()+" "+aylar[i]+" "+yıl)
                ws.cell(column=3,row=y,value=tahakkuk)
                ws.cell(column=5,row=y,value=tahakkuk)
                y += 1
                eski = son
                son = i
        öncekikes = 0
        if son!=0:
            if son!=aylar.index(self.aykombo.get()):
                eski = son
            öncekikes = ((self.parent.hesaplar[self.güzergahkombo.get()][eski][0]*self.parent.hesaplar[self.güzergahkombo.get()][eski][1])*
                         self.parent.hesaplar[self.güzergahkombo.get()][eski][3]+
                         (self.parent.hesaplar[self.güzergahkombo.get()][eski][0]*self.parent.hesaplar[self.güzergahkombo.get()][eski][1])*
                         self.parent.hesaplar[self.güzergahkombo.get()][eski][4])
        ws.cell(column=5,row=16,value=top)
        bukes = (hakediş*self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][3]+
                 hakediş*self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][4])
        ws.cell(column=5,row=19,value=bukes)
        ws.cell(column=5,row=20,value=öncekikes)
        ws.cell(column=5,row=21,value=bukes+öncekikes)
        ws.cell(column=5,row=23,value=top-(bukes+öncekikes))
        ws.cell(column=5,row=25,value=gettime())
        ws.cell(column=5,row=27,value=self.parent.ekbilgiler["ilçemüdür"])
        try:
            wb.save("icmalform.xlsx")
            os.startfile("icmalform.xlsx")
        except:
            messagebox.showerror("HATA","EXCEL DOSYASINI KAPATMADINIZ")
            
    def özet_yazdır(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA","GÜZERGAH SEÇ")
            return
        if self.aykombo.get()=="":
            messagebox.showerror("HATA","AY SEÇ")
            return
        if self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]=="":
            messagebox.showerror("HATA",f"{self.aykombo.get()} AYINA KAYIT YAPILMADI")
            return
        wb = pyx.load_workbook("Form.xlsx")
        for i in reversed(range(len(wb.worksheets))):
            if wb.sheetnames[i]!="ÖZET":
                wb.remove(wb.worksheets[i])
        ws = wb["ÖZET"]
        ws.cell(column=1,row=2,value=self.parent.bilgiler[self.güzergahkombo.get()][1]+" "+self.güzergahkombo.get())
        ws.cell(column=11,row=2,value=int(self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]))
        y = 6
        top = 0
        for i in range(aylar.index(self.aykombo.get())+1):
            if (self.parent.hesaplar[self.güzergahkombo.get()][i][0]!=0 or self.parent.hesaplar[self.güzergahkombo.get()][i][1]!=0):
                hakediş = self.parent.hesaplar[self.güzergahkombo.get()][i][0]*self.parent.hesaplar[self.güzergahkombo.get()][i][1]
                tahakkuk = hakediş+hakediş*self.parent.hesaplar[self.güzergahkombo.get()][i][2]
                top += tahakkuk
                yıl = self.parent.yıllar[self.güzergahkombo.get()][i][0]
                no = self.parent.yıllar[self.güzergahkombo.get()][i][1]
                ws.cell(column=1,row=y,value=yıl)
                ws.cell(column=2,row=y,value=int(no))
                ws.cell(column=3,row=y,value=top)
                ws.cell(column=11,row=y,value=tahakkuk)
                y += 1
        ws.cell(column=11,row=16,value=top)
        ws.cell(column=10,row=18,value=gettime())
        ws.cell(column=10,row=20,value=self.parent.ekbilgiler["ilçemüdür"])
        try:
            wb.save("icmalform.xlsx")
            os.startfile("icmalform.xlsx")
        except:
            messagebox.showerror("HATA","EXCEL DOSYASINI KAPATMADINIZ")
            
    def teklif_yazdır(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA","GÜZERGAH SEÇ")
            return
        if self.aykombo.get()=="":
            messagebox.showerror("HATA","AY SEÇ")
            return
        if self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]=="":
            messagebox.showerror("HATA",f"{self.aykombo.get()} AYINA KAYIT YAPILMADI")
            return
        wb = pyx.load_workbook("Form.xlsx")
        for i in reversed(range(len(wb.worksheets))):
            if wb.sheetnames[i]!="TEKLİF":
                wb.remove(wb.worksheets[i])
        ws = wb["TEKLİF"]
        ws.cell(column=6,row=3,value=self.güzergahkombo.get())
        ws.cell(column=6,row=4,value=self.parent.bilgiler[self.güzergahkombo.get()][1])
        ws.cell(column=6,row=5,value=self.parent.bilgiler[self.güzergahkombo.get()][9])
        ws.cell(column=6,row=6,value=self.parent.bilgiler[self.güzergahkombo.get()][5])
        ws.cell(column=6,row=7,value=self.parent.bilgiler[self.güzergahkombo.get()][4])
        ws.cell(column=6,row=8,value=self.parent.bilgiler[self.güzergahkombo.get()][7])
        ws.cell(column=6,row=10,value=self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
        ws.cell(column=6,row=11,value=self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
        metin1 = f"Öğrenci taşıma işinin bitirildiğine ilişkin yüklenici {self.parent.bilgiler[self.güzergahkombo.get()][1]},verdiği {self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]} tarihli dilekçe(fatura) üzerine İlçe Milli Eğitim Müdürünün talimatıile, yukarıda yazılı işin ön incelemesi {self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]} tarihinde tarafımızdan yapılmış, işin sözleşmesine uygun olarak tamamlandığı ve kabule hazır olduğu tespit edilmiştir.Gereğini arz ederim."
        ws.cell(column=2,row=13,value=metin1)
        ws.cell(column=10,row=16,value=self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
        ws.cell(column=9,row=18,value=self.parent.bilgiler[self.güzergahkombo.get()][1])
        ws.cell(column=2,row=29,value=self.parent.bilgiler[self.güzergahkombo.get()][12])
        ws.cell(column=6,row=29,value=self.parent.bilgiler[self.güzergahkombo.get()][13])
        ws.cell(column=10,row=29,value=self.parent.bilgiler[self.güzergahkombo.get()][14])
        ws.cell(column=5,row=40,value=gettime())
        ws.cell(column=5,row=41,value=self.parent.ekbilgiler["ilçemüdür"])
        ws.cell(column=6,row=56,value=self.güzergahkombo.get())
        ws.cell(column=6,row=57,value=self.parent.bilgiler[self.güzergahkombo.get()][1])
        ws.cell(column=6,row=58,value=self.parent.bilgiler[self.güzergahkombo.get()][9])
        ws.cell(column=6,row=59,value=self.parent.bilgiler[self.güzergahkombo.get()][5])
        ws.cell(column=6,row=60,value=self.parent.bilgiler[self.güzergahkombo.get()][4])
        ws.cell(column=6,row=61,value=self.parent.bilgiler[self.güzergahkombo.get()][7])
        ws.cell(column=6,row=63,value=self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
        ws.cell(column=6,row=64,value=self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
        metin2 = f"Müdürlüğümüz ile {self.parent.bilgiler[self.güzergahkombo.get()][1]}, arasında sözleşme kapsamında gerçekleştirilen iş için düzenlenen KABUL TEKLİF BELGESİ nden ön incelemenin yapıldığı anlaşılmış olup {self.olur.get()}  tarihli makam oluru ile Muayene ve kabul komisyonu, kontrol teşkilatı ve yüklenici de hazır olduğu halde taşınan ay için yüklenici tarafından yapılmış işleri kabul bakımından incelemiş ve aşağıda yazılı hususları tespit etmiştir ;"
        ws.cell(column=2,row=66,value=metin2)
        metin3 = f"SONUÇ: Kabul bakımından muayene ve inceleme işlemlerinin yapılması görevi Komisyonumuza verilmiş bulunan söz konusu işin yukarıda belirtilen (varsa ayrıntıları veya gerekçeleri ekli sayfalarda sayılan ve gösterilen)kayıtlarla ve bitim tarihi de  {self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]} olarak itibar edilmek üzere kabulünün yapılması Komisyonumuzca uygun görülmüş  ve İlçe Milli Eğitim Müdürlüğü Makamın onayına sunulmak üzere işbu Kabul Tutanağı 3 (üç) nüsha olarak düzenlenmiştir."
        ws.cell(column=2,row=68,value=metin3)
        ws.cell(column=10,row=70,value=self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
        ws.cell(column=9,row=72,value=self.parent.bilgiler[self.güzergahkombo.get()][1])
        ws.cell(column=2,row=81,value=self.parent.bilgiler[self.güzergahkombo.get()][12])
        ws.cell(column=6,row=81,value=self.parent.bilgiler[self.güzergahkombo.get()][13])
        ws.cell(column=10,row=81,value=self.parent.bilgiler[self.güzergahkombo.get()][14])
        ws.cell(column=5,row=93,value=gettime())
        ws.cell(column=5,row=94,value=self.parent.ekbilgiler["ilçemüdür"])
        try:
            wb.save("teklifform.xlsx")
            os.startfile("teklifform.xlsx")
        except:
            messagebox.showerror("HATA","EXCEL DOSYASINI KAPATMADINIZ")

    def teslim_yazdır(self):
        if self.güzergahkombo.get()=="":
            messagebox.showerror("HATA","GÜZERGAH SEÇ")
            return
        if self.aykombo.get()=="":
            messagebox.showerror("HATA","AY SEÇ")
            return
        if self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]=="":
            messagebox.showerror("HATA",f"{self.aykombo.get()} AYINA KAYIT YAPILMADI")
            return
        wb = pyx.load_workbook("Form.xlsx")
        for i in reversed(range(len(wb.worksheets))):
            if wb.sheetnames[i]!="TESLİM":
                wb.remove(wb.worksheets[i])
        ws = wb["TESLİM"]
        ws.cell(column=19,row=5,value=self.birim.get())
        ws.cell(column=3,row=6,value=self.daire.get())
        an = datetime.now()
        string = ""
        if an.day<10:
            string += "0"
        string += str(an.day)+"."
        if an.month<10:
            string += "0"
        string += str(an.month)+"."
        string += str(an.year)
        ws.cell(column=19,row=6,value=string)
        ws.cell(column=20,row=11,value=self.parent.bilgiler[self.güzergahkombo.get()][1])
        ws.cell(column=23,row=12,value=self.parent.bilgiler[self.güzergahkombo.get()][2])
        hakediş = self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0]*self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][1]
        tahakkuk = hakediş*self.parent.hesaplar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][2]+hakediş
        ws.cell(column=24,row=12,value=tahakkuk)
        ws.cell(column=20,row=13,value=self.parent.yıllar[self.güzergahkombo.get()][aylar.index(self.aykombo.get())][0])
        ws.cell(column=20,row=14,value=self.güzergahkombo.get())
        metin1 = f"Yukarıda alacaklıları ile alacak tutarları gösterilen toplam ... adet tahukkuk evrağı ve ekleri teslim alınmıştır. {gettime()}  Teslim Saati ....... / ..........  )"
        ws.cell(column=2,row=26,value=metin1)
        ws.cell(column=4,row=31,value=self.teslim.get())
        try:
            wb.save("teslimform.xlsx")
            os.startfile("teslimform.xlsx")
        except:
            messagebox.showerror("HATA","EXCEL DOSYASINI KAPATMADINIZ")
        
class main:
    def __init__(self):
        self.active = 1
        self.bilgiler = {}
        self.ekbilgiler = {"müdür":"","ilçemüdür":"","teslim":"","daire":"","birim":"","olur":""}
        self.yıllar = {}
        self.hesaplar = {}
        self.load_data()
        self.pencereler = {"bilgi":bilgi_pencere(self),"hakediş":hakediş(self),"icmal":icmal(self)}
        self.pencere = Tk()
        self.pencere.resizable(0,0)
        self.pencere.title("MENÜ")
        self.canvas = Canvas(self.pencere,width=300,height=250,highlightthickness=0,bg="#0D252B")
        self.canvas.pack()
        self.canvas.update()
        self.resx = self.canvas.winfo_width()
        self.resy = self.canvas.winfo_height()
        buton = Button(self.canvas,text="BİLGİ GİR",width=25,command=self.pencereler["bilgi"].show)
        self.canvas.create_window(self.resx/2,70,window=buton)
        buton = Button(self.canvas,text="HAKEDİŞ HESAPLA",width=25,command=self.pencereler["hakediş"].show)
        self.canvas.create_window(self.resx/2,120,window=buton)
        buton = Button(self.canvas,text="TUTANAKLAR",width=25,command=self.pencereler["icmal"].show)
        self.canvas.create_window(self.resx/2,170,window=buton)

    def update(self):
        self.pencere.update()
        self.pencereler["bilgi"].update()
        self.pencereler["hakediş"].update()
        self.pencereler["icmal"].update()
        
    def show(self):
        self.pencere.deiconify()
        
    def hide(self):
        self.pencere.withdraw()
        
    def load_data(self):
        try:
            file = np.load("data.npz",allow_pickle=True)
        except:
            self.save_data()
            return
        
        self.bilgiler = file["bilgiler"].tolist()
        self.hesaplar = file["hesaplar"].tolist()
        self.ekbilgiler = file["müdürler"].tolist()
        self.yıllar = file["yıllar"].tolist()
    def save_data(self):
        file = {}
        file["bilgiler"] = self.bilgiler
        file["hesaplar"] = self.hesaplar
        file["müdürler"] = self.ekbilgiler
        file["yıllar"] = self.yıllar
        np.savez("data.npz",**file)

def trunc(x):
    string = str(x)
    if "." in string:
        string1 = string[:string.index(".")]
        string2 = string[string.index("."):]
        if len(string2)<3:
            return string1+string2+(len(string2)-1)*"0"
        return string1+string2[:3]
    return string+".00"

harf = "1234567890.,"
def safe(x):
    string = str(x)
    if string.count(".")+string.count(",")>1:
        return -1
    for index,i in enumerate(string):
        if not i in harf:
            return -1
        if i==",":
            yeni = string[:index]+"."
            if index!=len(string)-1:
                yeni += string[index+1:]
            string = yeni
    return string

def gettime():
    an = datetime.now()
    string = "..... / "
    if an.month<10:
        string += "0"
    string += str(an.month)
    string += " / "
    string += str(an.year)
    return string

pencere = main()
while 1:
    try:
        pencere.update()
    except:
        break
    time.sleep(1/60)
