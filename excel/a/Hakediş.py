from openpyxl import *
from copy import deepcopy
from tkinter import *
from tkinter.messagebox import *
aylar = {9:4,
         10:7,
         11:10,
         12:13,
         0o1:16,
         0o2:19,
         0o3:22,
         0o4:25,
         0o5:28,
         0o6:31}
yerler = {1:"A",
          4:"D",
          7:"G",
          10:"J",
          13:"M",
          16:"P",
          19:"S",
          22:"V",
          25:"Y",
          28:"AB",
          31:"AE"}
def writeto(ws,row,col,value,font=0):
    wc = ws.cell(row,col)
    wc.value = value
    if font!=0:
        ws.font = font
def loadfrom(ws,name):
    return ws[name].value
başlama = 0
def hakediş(kişi_sayısı):
    global başlama
    if kişi_sayısı != "":
        try:
            kişi_sayısı = int(kişi_sayısı)
        except:
            showerror("Hata","Kişi Sayısını Hatalı Girdiniz")
        else:
            wb = load_workbook(filename='..\Veri.xlsx', read_only=True)
            ws = wb['Sayfa1']
            isimler = []
            no = loadfrom(ws,"b4")
            tarih = loadfrom(ws,"c4")
            eski = deepcopy(tarih)
            try:
                new = ""+str(tarih.day)+"."+str(tarih.month)+"."+str(tarih.year)
                tarih = new
            except:
                pass
            günler = []
            fiyatlar = []
            hakedişler = []
            kdv_oranlar = []
            kdv_tutarlar = []
            kdv_tevkler = []
            damgalar = []
            tahakkuklar = []
            kesinti_toplar = []
            ödenecekler = []
            for i in range(4,4+kişi_sayısı):
                isimler.append(loadfrom(ws,f"A{i}"))
                günler.append(loadfrom(ws,f"d{i}"))
                fiyatlar.append(loadfrom(ws,f"e{i}"))
                hakedişler.append(günler[-1]*fiyatlar[-1])
                kdv_oranlar.append(float(loadfrom(ws,f"g{i}")))
                kdv_tutarlar.append(hakedişler[-1]*kdv_oranlar[-1])
                kdv_tevkler.append(kdv_tutarlar[-1]/2)
                damgalar.append(round(hakedişler[-1]*0.00949,2))
                tahakkuklar.append(hakedişler[-1]+kdv_tutarlar[-1])
                kesinti_toplar.append(kdv_tevkler[-1]+damgalar[-1])
                ödenecekler.append(tahakkuklar[-1]-kesinti_toplar[-1])
            wb.close()
            wb = load_workbook(filename='..\Yedekler\Veri.xlsx')
            ws = wb["Sayfa1"]
            for i in range(0,len(isimler)):
                writeto(ws,i+4,1,isimler[i])
                writeto(ws,i+4,2,no)
                writeto(ws,i+4,3,tarih)
                writeto(ws,i+4,4,günler[i])
                writeto(ws,i+4,5,fiyatlar[i])
                writeto(ws,i+4,6,hakedişler[i])
                writeto(ws,i+4,7,kdv_oranlar[i])
                writeto(ws,i+4,8,kdv_tutarlar[i])
                writeto(ws,i+4,9,kdv_tevkler[i])
                writeto(ws,i+4,10,damgalar[i])
                writeto(ws,i+4,11,tahakkuklar[i])
                writeto(ws,i+4,12,kesinti_toplar[i])
                writeto(ws,i+4,13,ödenecekler[i])
            wb.save("..\Hazırlananlar\Girilen_Veriler.xlsx")
            wb = load_workbook(filename='..\Hazırlananlar\Aylar.xlsx')
            ws = wb["Sayfa1"]
            for i in range(0,len(isimler)):
                try:
                    writeto(ws,i+4,aylar[eski.month],hakedişler[i])
                except:
                    writeto(ws,i+4,aylar[int(str(eski)[3:5])],hakedişler[i])
            wb.save("..\Hazırlananlar\Aylar.xlsx")
            toplamlar = []
            wb = load_workbook(filename='..\Hazırlananlar\Aylar.xlsx')
            ws = wb["Sayfa1"]
            for i in range(0,len(isimler)):
                toplam = 0
                for y in range(9,13):
                    toplam += float(loadfrom(ws,f"{yerler[aylar[y]]}{i+4}"))
                for y in range(1,7):
                    toplam += float(loadfrom(ws,f"{yerler[aylar[y]]}{i+4}"))
                toplamlar.append(toplam)
            wb.close()
            wb = load_workbook(filename='..\Yedekler\Hakediş.xlsx')
            for i in range(0,len(isimler)):
                try:
                    ws = wb[isimler[i]]
                except:
                    showerror("Yanlış isim",f"{isimler[i]} adlı kişi Yedekler\Hakediş.xlsx adlı dosyada bulunmuyor")
                    quit()
                writeto(ws,5,4,toplamlar[i])
                writeto(ws,8,4,toplamlar[i])
                writeto(ws,2,6,no)
                writeto(ws,10,4,toplamlar[i]-hakedişler[i])
                writeto(ws,4,1,tarih)
                writeto(ws,11,4,hakedişler[i])
                writeto(ws,12,4,kdv_tutarlar[i])
                writeto(ws,13,4,tahakkuklar[i])
                writeto(ws,15,4,damgalar[i])
                writeto(ws,16,4,kdv_tevkler[i])
                writeto(ws,24,4,kesinti_toplar[i])
                writeto(ws,25,4,ödenecekler[i])
                writeto(ws,28,2,isimler[i])
            wb.save("..\Hazırlananlar\Hakediş.xlsx")
            showinfo("Başarılı","Başarılı Artık Kapatabilirsiniz")
    elif başlama==1:
        showerror("Hata","Kişi Sayısı Boş Bırakılamaz")
pencere = Tk()
pencere.title("Hakediş")
label = Label(pencere,text="Kişi Sayısı")
label.grid(row=0,column=0)
e = Entry(pencere,width=30)
e.grid(row=0,column=1,columnspan=3)
düğme = Button(pencere,text="Hazırla",padx=30,pady=10,command=lambda : hakediş(e.get()))
düğme.grid(row=1,column=1)
başlama = 1
pencere.mainloop()
