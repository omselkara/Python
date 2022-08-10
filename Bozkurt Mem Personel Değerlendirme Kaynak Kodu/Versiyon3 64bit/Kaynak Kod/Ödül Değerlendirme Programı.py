from tkinter import *
from tkinter.ttk import Combobox
import time
from tkinter import filedialog as fd
import tkinter
import openpyxl as pyx
from openpyxl.styles.borders import Border, Side
from tkinter import messagebox
import os
import datetime

num_range = 4

def turn_readable(n):
    return "0"*(num_range-len(str(n)))+str(n)
    
def load(name):
    file = open(name,"r",encoding="utf-8")
    file = file.readlines()
    liste = []
    for i in file:
        if "\n" in i:
            liste.append(i[0:-1])
        else:
            liste.append(i)
    text = ""
    for i in liste:
        text += i
    file = decode(text)
    return eval(f"{file}")

def decode(file):
    text = ""
    for s in range(0,len(file),num_range):
        text += chr(int(file[s:s+num_range]))
    return text

def encode(texts):
    liste = []
    text = ""
    for s in texts:
        text += turn_readable(ord(s))
        if len(text)==200:
            liste.append(text)
            text = ""
    if text!="":
        liste.append(text)
    return liste

def save(name,texts):
    file = open(name,"w",encoding="utf-8")
    texts = encode(str(texts))
    for i in texts:
        file.write(i+"\n")
    file.close()


class kişi_kayıt:
    def __init__(self,index):
        self.bilgiler = []
        self.puanlar = [[],[],[],[]]
        self.index = index

class ana_pencere:
    def __init__(self):
        self.açık = True
        self.pencere = Tk()
        self.pencere.resizable(0,0)
        self.pencere.title("Ödül Değerlendirme Programı")
        self.pencere.geometry("+300+0")
        self.canvas = Canvas(self.pencere,width=300,height=480,highlightthickness=0,bg="#424744")
        self.canvas.pack()
        self.canvas.create_text(150,460,text="Ömer Selim KARAŞAH",font=("italic",8,"bold"),fill="#555555")
        self.ödül_alacak = Button(self.canvas,text="Bilgi Girişi",command=lambda :self.basıldı("ödül pencere"))
        self.ödül_alacak.place(x=55,y=45,width=190,height=40)
        self.disiplin_amir = Button(self.canvas,text="Disiplin Amiri Değerlendirme\n Formu",command=lambda :self.basıldı("disiplin pencere"))
        self.disiplin_amir.place(x=55,y=130,width=190,height=40)#(x=360,y=50,width=190,height=40)
        self.ilçe_değerlendirme = Button(self.canvas,text="İlçe Değerlendirme Komisyon\n Formu",command=lambda :self.basıldı("ilçe pencere"))
        self.ilçe_değerlendirme.place(x=55,y=215,width=190,height=40)#(x=50,y=150,width=190,height=40)
        self.il_değerlendirme = Button(self.canvas,text="İl'e Gönderilecek Bilgi Formu Yazdır",command=lambda :self.basıldı("il yazdır"))
        self.il_değerlendirme.place(x=55,y=300,width=190,height=40)#(x=50,y=150,width=190,height=40)
        self.il_değerlendirme = Button(self.canvas,text="İl Ödül Değerlendirme İşlemi",command=lambda :self.basıldı("il form yazdır"),bg="Gray")
        self.il_değerlendirme.place(x=55,y=385,width=190,height=40)#(x=50,y=150,width=190,height=40)
        self.pencereler = {"ödül pencere":ödül_pencere(self),"disiplin pencere":disiplin_pencere(self),"ilçe pencere":ilçe_pencere(self),"hakkında pencere":hakkında_pencere(self)}
        self.pencere.protocol("WM_DELETE_WINDOW", self.kapat)
        self.pencereler["disiplin pencere"].update()
        self.pencereler["ilçe pencere"].update()
        self.pencereler["hakkında pencere"].update()
        self.pencere.iconbitmap('Bayrak.ico')


    def basıldı(self,isim):
        if isim=="il yazdır" or isim=="il form yazdır":
            if len(self.pencereler["ödül pencere"].kişiler)>0:
                thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
                wb = pyx.load_workbook("İl Form.xlsx")
                ws = wb.active
                sıralama = []
                yazılcaklar = []
                for num,kişi in enumerate(self.pencereler["ödül pencere"].kişiler):
                    üyeler = []
                    unvanları = []
                    indexler = []
                    for index,i in enumerate(kişi.puanlar[1]):
                        boş = True
                        for y in i:
                            if y!="":
                                try:
                                    d = int(y)
                                except:
                                    continue
                                if d>0:
                                    boş = False
                                    break
                        if not boş:
                            üyeler.append(self.pencereler["ödül pencere"].komisyon[index])
                            unvanları.append(self.pencereler["ödül pencere"].unvanlar[index])
                            indexler.append(index)
                    hepsi = 0
                    for i in range(13,46):
                        toplam = 0
                        sayı = 0
                        for index in indexler:
                            if str(kişi.puanlar[1][index][i-13])!="":
                                try:
                                    val = int(kişi.puanlar[1][index][i-13])
                                except:
                                    continue
                                sayı += 1
                                toplam += val
                        if sayı!=0:
                            toplam = toplam/sayı
                        sayı = 0
                        val2 = 0
                        if toplam!=0:
                            sayı += 1
                        if i-13<len(kişi.puanlar[0]) and kişi.puanlar[0][i-13]!="":
                            try:
                                d = int(kişi.puanlar[0][i-13])
                                val2 = d
                            except:
                                pass
                            if val2!=0:
                                sayı+=1
                        if sayı!=0:
                            hepsi += (toplam+val2)/sayı
                    sıralama.append(hepsi)
                    yazılcaklar.append([num+1,kişi.bilgiler[1],kişi.bilgiler[0],kişi.bilgiler[7],kişi.bilgiler[8],
                    kişi.bilgiler[4],kişi.bilgiler[3],hepsi,kişi.bilgiler[5],kişi.bilgiler[9],kişi.bilgiler[6]])
                    #ws.cell(column=1,row=3+num).border = thin_border
                    #ws.cell(column=1,row=3+num).value = num+1
                    #ws.cell(column=2,row=3+num).border = thin_border
                    #ws.cell(column=2,row=3+num).value = kişi.bilgiler[1]
                    #ws.cell(column=3,row=3+num).border = thin_border
                    #ws.cell(column=3,row=3+num).value = kişi.bilgiler[0]
                    #ws.cell(column=4,row=3+num).border = thin_border
                    #ws.cell(column=4,row=3+num).value = kişi.bilgiler[7]
                    #ws.cell(column=5,row=3+num).border = thin_border
                    #ws.cell(column=5,row=3+num).value = kişi.bilgiler[8]
                    #ws.cell(column=6,row=3+num).border = thin_border
                    #ws.cell(column=6,row=3+num).value = kişi.bilgiler[4]
                    #ws.cell(column=7,row=3+num).border = thin_border
                    #ws.cell(column=7,row=3+num).value = kişi.bilgiler[3]
                    #ws.cell(column=8,row=3+num).border = thin_border
                    #ws.cell(column=8,row=3+num).value = hepsi
                    #ws.cell(column=9,row=3+num).border = thin_border
                    #ws.cell(column=9,row=3+num).value = kişi.bilgiler[5]
                    #ws.cell(column=10,row=3+num).border = thin_border
                    #ws.cell(column=10,row=3+num).value = kişi.bilgiler[9]
                    #ws.cell(column=11,row=3+num).border = thin_border
                    #ws.cell(column=11,row=3+num).value = kişi.bilgiler[6]
                for _ in range(len(sıralama)):
                    best = -10000
                    index = 0
                    if isim=="il form yazdır":
                        for i in range(0,len(sıralama)):
                            if sıralama[i]>best:
                                best = sıralama[i]
                                index = i
                    ws.cell(column=1,row=3+_).border = thin_border
                    ws.cell(column=1,row=3+_).value = yazılcaklar[index][0]
                    ws.cell(column=2,row=3+_).border = thin_border
                    ws.cell(column=2,row=3+_).value = yazılcaklar[index][1]
                    ws.cell(column=3,row=3+_).border = thin_border
                    ws.cell(column=3,row=3+_).value = yazılcaklar[index][2]
                    ws.cell(column=4,row=3+_).border = thin_border
                    ws.cell(column=4,row=3+_).value = yazılcaklar[index][3]
                    ws.cell(column=5,row=3+_).border = thin_border
                    ws.cell(column=5,row=3+_).value = yazılcaklar[index][4]
                    ws.cell(column=6,row=3+_).border = thin_border
                    ws.cell(column=6,row=3+_).value = yazılcaklar[index][5]
                    ws.cell(column=7,row=3+_).border = thin_border
                    ws.cell(column=7,row=3+_).value = yazılcaklar[index][6]
                    ws.cell(column=8,row=3+_).border = thin_border
                    ws.cell(column=8,row=3+_).value = yazılcaklar[index][7]
                    ws.cell(column=9,row=3+_).border = thin_border
                    ws.cell(column=9,row=3+_).value = yazılcaklar[index][8]
                    ws.cell(column=10,row=3+_).border = thin_border
                    ws.cell(column=10,row=3+_).value = yazılcaklar[index][9]
                    ws.cell(column=11,row=3+_).border = thin_border
                    ws.cell(column=11,row=3+_).value = yazılcaklar[index][10]
                    sıralama.pop(index)
                    yazılcaklar.pop(index)
                il = str(kişi.bilgiler[7]).replace("i","İ").upper()
                yıl = str(datetime.datetime.now().year)
                ws.cell(column=1,row=1).value = str(il)+" - "+str(yıl)+" YILINDA ÖDÜL ALACAK PERSONELE DAİR BİLGİLER"
                if isim=="il yazdır":
                    filename = fd.asksaveasfilename(initialfile=f"İl Bilgi Form.xlsx",defaultextension=".xlsx",filetypes=[("Excel Files","*.xlsx")])
                else:
                    filename = fd.asksaveasfilename(initialfile=f"İl Ödül Değerlendirme Form.xlsx",defaultextension=".xlsx",filetypes=[("Excel Files","*.xlsx")])
                if filename=="":
                    return
                try:
                    wb.save(filename)
                    os.startfile(filename)
                except:
                    messagebox.showerror("Hata","Excel Dosyasını Kapatmadınız")
        else:
            self.pencere.withdraw()
            self.pencereler[isim].pencere.deiconify()
        
    def show(self):
        self.canvas.update()
        self.pencere.update()
        for i in self.pencereler:
            self.pencereler[i].show()

    def kapat(self):
        self.açık = False
        self.pencere.destroy()
        self.pencere.quit()

class ödül_pencere:
    def __init__(self,parent):
        self.eski_seçili_index = 0
        self.kişiler = []
        self.komisyon = []
        self.unvanlar = []
        self.parent = parent
        self.pencere = Tk()
        self.pencere.resizable(0,0)
        self.pencere.geometry("+300+0")
        self.pencere.title("Ödül Değerlendirme Programı")
        self.width = 900
        self.height = 500
        self.canvas = Canvas(self.pencere,width=self.width,height=self.height,highlightthickness=0,bg="#424744")
        self.canvas.pack()
        self.pencere.withdraw()
        self.pencere.iconbitmap('Bayrak.ico')
        self.pencere.protocol("WM_DELETE_WINDOW", self.kapat)
        self.isimler = ["Adı Soyadı","TC No","Kurum Sicil No","Kadro Unvanı Ve Görevi","Görev Yeri Ve Kurumu",
        "Kazanılmış Hak Aylıkta Değerlendirilen Hizmet Süresi","Üstün Başarı Belgesi Tarih Ve Sayısı","İl","İlçe","Yıl İçerisinde Kınama Cezasından Ağır Ceza Almışmı",
                        "Disiplin Amirinin Adı","Disiplin Amirinin Unvanı","","",""]
        self.entryler = []
        for y in range(len(self.isimler)):
            self.canvas.create_rectangle(0,y*(self.height-100)/len(self.isimler),300,(y+1)*(self.height-100)/len(self.isimler),fill="#AAAAAA")
            self.canvas.create_text(10,(y*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text=self.isimler[y],anchor=W)
            entry = Entry(self.canvas)
            entry.place(x=300,y=y*(self.height-100)/len(self.isimler)+(self.height-100)/(len(self.isimler)*10),width=350,height=(self.height-100)/(len(self.isimler)*10/8))
            self.entryler.append(entry)
        self.seçme = Combobox(self.canvas)
        self.seçme["state"] = "readonly"
        self.seçme["values"] = [""]+[i.bilgiler[0] for i in self.kişiler]
        self.seçme.bind("<<ComboboxSelected>>",self.yeni_seçildi)
        self.seçme.place(x=670,y=10,width=210)
        self.kaydet_button = Button(self.canvas,text="Kaydet",command=self.kaydet_basıldı)
        self.kaydet_button.place(x=730,y=70,width=100,height=30)
        self.sil_button = Button(self.canvas,text="Seçili Kişiyi Sil",command=self.sil_basıldı)
        self.sil_button.place(x=730,y=150,width=100,height=30)
        self.tümünü_sil_button = Button(self.canvas,text="Tümünü Sil",command=self.tümünü_sil_basıldı)
        self.tümünü_sil_button.place(x=730,y=230,width=100,height=30)
        self.yedekle_button = Button(self.canvas,text="Yedekle",command=self.yedekle_basıldı)
        self.yedekle_button.place(x=730,y=310,width=100,height=30)
        self.yedekten_al_button = Button(self.canvas,text="Yedekten Al",command=self.yükle_basıldı)
        self.yedekten_al_button.place(x=730,y=390,width=100,height=30)
        self.yedekten_ekle_button = Button(self.canvas,text="(İl İlçe) Verileri\n Birleştir",command=self.yedekten_ekle_basıldı,bg="Gray")
        self.yedekten_ekle_button.place(x=730,y=460,width=100,height=34)
        self.kayıt_sayı_yazı = self.canvas.create_text(700,50,text="Kayıt Sayısı:0",fill="white")
        self.excelden_ekle_button = Button(self.canvas,text="Excelden Kişiler\n Al",command=self.excelden_ekle_basıldı,bg="Gray")
        self.excelden_ekle_button.place(x=70,y=460,width=100,height=34)
        self.hakkında = Button(self.canvas,text="Hakkında",command=self.hakkında_basıldı)
        self.hakkında.place(x=4,y=410,width=53,height=17)
        self.load()

    def hakkında_basıldı(self):
        self.pencere.withdraw()
        self.parent.pencereler["hakkında pencere"].pencere.deiconify()
        
    def excelden_ekle_basıldı(self):
        filename = fd.askopenfilename(initialfile="",defaultextension=".xlsx",filetypes=[("Excel Files","*.xlsx")])
        if filename=="":
            return
        wb = pyx.load_workbook(filename)
        ws = wb.active
        y = 3
        while 1:
            if ws.cell(column=1,row=y).value==None:
                break
            kişi = kişi_kayıt(len(self.kişiler))
            for s in range(len(self.komisyon)):
                kişi.puanlar[1].append([])
                kişi.puanlar[2].append([])
                kişi.puanlar[3].append([])
            if ws.cell(column=3,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=3,row=y).value))
            if ws.cell(column=2,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=2,row=y).value))
            if ws.cell(column=12,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=12,row=y).value))
            if ws.cell(column=7,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=7,row=y).value))
            if ws.cell(column=6,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=6,row=y).value))
            if ws.cell(column=9,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=9,row=y).value))
            if ws.cell(column=11,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=11,row=y).value))
            if ws.cell(column=4,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=4,row=y).value))
            if ws.cell(column=5,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=5,row=y).value))
            if ws.cell(column=10,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=10,row=y).value))
            if ws.cell(column=13,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=13,row=y).value))
            if ws.cell(column=14,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=14,row=y).value))
            if ws.cell(column=15,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=15,row=y).value))
            if ws.cell(column=16,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=16,row=y).value))
            if ws.cell(column=17,row=y).value==None:
                kişi.bilgiler.append("")
            else:
                kişi.bilgiler.append(str(ws.cell(column=17,row=y).value))
            self.kişiler.append(kişi)
            y += 1
        self.update()
        self.kayıt_yap()
        self.parent.pencereler["disiplin pencere"].update()
        self.parent.pencereler["ilçe pencere"].update()

    def yedekle_basıldı(self):
        filename = fd.asksaveasfilename(initialfile="Yedek.txt",defaultextension=".txt",filetypes=[("Txt Files","*.txt")])
        if filename=="":
            return
        schema = [[i.bilgiler for i in self.kişiler],[i.puanlar for i in self.kişiler],self.komisyon,self.unvanlar]
        save(filename,schema)
    
    def yükle_basıldı(self):
        filename = fd.askopenfilename(initialfile="Yedek.txt",defaultextension=".txt",filetypes=[("txt Files","*.txt")])
        if filename=="":
            return
        schema = load(filename)
        self.kişiler = []
        self.komisyon = [i for i in schema[2]]
        self.unvanlar = [i for i in schema[3]]
        for i in range(len(schema[0])):
            kişi = kişi_kayıt(len(self.kişiler))
            kişi.bilgiler = schema[0][i]
            kişi.puanlar = schema[1][i]
            self.kişiler.append(kişi)
        self.seçme.current(0)
        self.yeni_seçildi(0)
        self.update()
        self.kayıt_yap()
        self.parent.pencereler["disiplin pencere"].update()
        self.parent.pencereler["ilçe pencere"].update()

    def yedekten_ekle_basıldı(self):
        filename = fd.askopenfilename(initialfile="Yedek.txt",defaultextension=".txt",filetypes=[("txt Files","*.txt")])
        if filename=="":
            return
        schema = load(filename)
        #self.komisyon = schema[2]
        #self.unvanlar = [i for i in schema[3]]
        for i in range(len(schema[0])):
            kişi = kişi_kayıt(len(self.kişiler))
            kişi.bilgiler = schema[0][i]
            kişi.puanlar = schema[1][i]
            self.kişiler.append(kişi)
        self.seçme.current(0)
        self.yeni_seçildi(0)
        self.update()
        self.kayıt_yap()
        self.parent.pencereler["disiplin pencere"].update()
        self.parent.pencereler["ilçe pencere"].update()

    def yeni_seçildi(self,event):
        if self.seçme.current()!=self.eski_seçili_index:
            self.eski_seçili_index = self.seçme.current()
            for i in range(0,len(self.entryler)):
                self.entryler[i].delete(0,END)
                if self.seçme.current()!=0:
                    self.entryler[i].insert(0,self.kişiler[self.seçme.current()-1].bilgiler[i])
    
    def kaydet_basıldı(self):
        if self.entryler[0].get()=="":
            return
        if self.seçme.current()==0:
            kişi = kişi_kayıt(len(self.kişiler))
            for y in range(len(self.komisyon)):
                kişi.puanlar[1].append([])
                kişi.puanlar[2].append([])
                kişi.puanlar[3].append([])
            for i in self.entryler:
                kişi.bilgiler.append(i.get())
                i.delete(0,END)
            self.kişiler.append(kişi)
        else:
            kişi = self.kişiler[self.seçme.current()-1]
            kişi.bilgiler = []
            for i in self.entryler:
                kişi.bilgiler.append(i.get())
                i.delete(0,END)
        self.update()
        self.kayıt_yap()
        self.parent.pencereler["disiplin pencere"].update()
        self.parent.pencereler["ilçe pencere"].update()
        self.seçme.current(0)
        self.eski_seçili_index = 0

    def sil_basıldı(self):
        if self.seçme.current()!=0:
            answer = messagebox.askquestion("Eminmisiniz","Eminmisiniz")
            if answer=="yes":
                self.kişiler.pop(self.seçme.current()-1)
                self.seçme.current(0)
                self.yeni_seçildi(0)
                self.update()
                self.kayıt_yap()
                self.parent.pencereler["disiplin pencere"].update()
                self.parent.pencereler["ilçe pencere"].update()
    
    def tümünü_sil_basıldı(self):
        answer = messagebox.askquestion("Eminmisiniz","Eminmisiniz")
        if answer=="yes":
            self.kişiler = []
            self.seçme.current(0)
            self.yeni_seçildi(0)
            self.update()
            self.kayıt_yap()
            self.parent.pencereler["disiplin pencere"].update()
            self.parent.pencereler["ilçe pencere"].update()

    def kapat(self):
        self.pencere.withdraw()
        self.parent.pencere.deiconify()
        self.seçme.current(0)
        for i in self.entryler:
            i.delete(0,END)

    def load(self):
        try:
            file = open("Yedek Ödül.txt","rb")
        except FileNotFoundError:
            schema = [[],[],[],[]]
            save("Yedek Ödül.txt",schema)
            return 
        schema = load("Yedek Ödül.txt")
        self.kişiler = []
        self.komisyon = [i for i in schema[2]]
        self.unvanlar = [i for i in schema[3]]
        for i in range(len(schema[0])):
            kişi = kişi_kayıt(len(self.kişiler))
            kişi.bilgiler = schema[0][i]
            kişi.puanlar = schema[1][i]
            self.kişiler.append(kişi)
        self.seçme.current(0)
        self.yeni_seçildi(0)
        self.update()

    def kayıt_yap(self):
        schema = [[i.bilgiler for i in self.kişiler],[i.puanlar for i in self.kişiler],self.komisyon,self.unvanlar]
        save("Yedek Ödül.txt",schema)

    def show(self):
        self.canvas.update()
        self.pencere.update()

    def update(self):
        self.seçme["values"] = [""]+[i.bilgiler[0] for i in self.kişiler]
        self.canvas.itemconfig(self.kayıt_sayı_yazı,text=f"Kayıt Sayısı:{len(self.kişiler)}")
        self.show()

class disiplin_pencere:
    def __init__(self,parent):
        self.parent = parent
        self.pencere = Tk()
        self.pencere.resizable(0,0)
        self.pencere.geometry("+300+0")
        self.pencere.title("Ödül Değerlendirme Programı")
        self.width = 900
        self.height = 650
        self.canvas = Canvas(self.pencere,width=self.width,height=self.height,highlightthickness=0,bg="#424744")
        self.canvas.pack()
        self.pencere.withdraw()
        self.pencere.iconbitmap('Bayrak.ico')
        self.pencere.protocol("WM_DELETE_WINDOW", self.kapat)
        self.isimler = ["Genel Nitelikler","Çevresi ile ilişkileri","Görevin gerektirdiği sır saklama ve gizliliğe riayeti",
                        "Görevinde tarafsızlık ve eşitlik ilkesine bağlılığı","Görevini benimseme ve sorumluluk alma özelliği","Genel kültür bilgisi",
                        "Görevini yapmadaki hevesi ve kendini yetiştirme gayreti","Bireysel iş görme kabiliyeti ve teşebbüs fikri",
                        "Meslek şeref ve haysiyetini zedeleyici kötü alışkanlıklara düşkün olmaması","Mesleki Yeterlilik","Mesleki bilgisi ve uygulamaya yansıtması",
                        "Görevini yapmadaki başarısı","Göreve başlamadan önce hazırlık yapma becerisi","Görevinin gerektirdiği faaliyetleri gerçekleştirmesi",
                        "Kullandığı eşyanın bakım, koruma, tertip ve düzenini sağlama","İşbirliği yapmada ve değişen şartlara, görevlere uymada gösterdiği başarı",
                        "Mevzuat Bilgisi, bağlılığı ve uygulaması","Mesleki, sosyal ve kültürel faliyet durumu",
                        "Görevi sırasında ortaya çıkan problemleri pratik ve akılcı bir şekilde çözümlemedeki başarısı",
                        "Uzun süreli çalışmalarda fikri ve bedeni dayanıklılığı","Mesleği ile İlgili Faaliyetleri, Olağanüstü Gayret ve Çalışmaları ile Diğer Hususlar",
                        "Gazete ve dergilerde yayımlanmış bilimsel ve mesleki inceleme,araştırma ve makaleler İçin","Yazmış olduğu (ISBN almış) kitapları için",
                        "Türkiye genelinde yapılan yarışmalarda;","a. Birinci olanlara","b. İkinci olanlara","c. Üçüncü olanlara","d. Mansiyon alanlara",
                        "Yurt dışı yarışmaların her birinde; ","a. Birinci olanlara","b. İkinci olanlara","c. Üçüncü olanlara","d. Mansiyon alanlara",
                        """Mesleği ile ilgili bir buluş yaparak veya mevcut usullerde yararlı
yenilikler meydana getirerek veyahut başka yolarla gelir ve verimliliğin arttırılmasında yahut
hizmetin yürütülmesinde dikkate değer çalışmalar yapmak (tescil edilmiş olmalı)"""]
        self.azamiler = ["Azami Puan",5,5,5,5,5,5,5,5,"",5,5,5,5,5,5,5,5,5,5,"",6,8,"",8,6,4,2,"",10,8,6,4,10,""]
        self.entryler = []
        self.engel = [9,20,23,28,34]
        self.canvas.create_rectangle(0,0,675,self.height-50,fill="#AAAAAA")
        self.canvas.create_text(624,7,text="Puan",anchor=W)
        for y in range(0,len(self.isimler)):
            self.canvas.create_rectangle(0,y*(self.height-100)/len(self.isimler),600,(y+1)*(self.height-100)/len(self.isimler),fill="#AAAAAA")
            if y!=33:
                self.canvas.create_rectangle(599,y*(self.height-100)/len(self.isimler),675,(y+1)*(self.height-100)/len(self.isimler),fill="#AAAAAA")
            if y!=len(self.isimler)-1:
                self.canvas.create_text(10,(y*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text=str(self.isimler[y]),#+f" (Azami Puan:{str(self.azamiler[y])})"
                                anchor=W)
                self.canvas.create_text(550,(y*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text=str(self.azamiler[y]),#+f" (Azami Puan:{str(self.azamiler[y])})"
                )
                if y!=0:
                    entry = Entry(self.canvas)
                    entry.bind("<KeyRelease>", self.değişim)
                    if not y in self.engel:
                        entry.place(x=600,y=y*(self.height-100)/len(self.isimler)+(self.height-100)/(len(self.isimler)*10),width=75,height=(self.height-100)/(len(self.isimler)*10/8))
                    self.entryler.append(entry)
        entry = Entry(self.canvas)
        entry.bind("<KeyRelease>", self.değişim)
        entry.place(x=600,y=(y)*(self.height-100)/len(self.isimler)+(self.height-100)/(len(self.isimler)*10),width=75,height=((self.height-100)/(len(self.isimler)*10/8))*3.5)
        self.entryler.append(entry)
        entry = Entry(self.canvas)
        entry.bind("<KeyRelease>", self.değişim)
        entry.place(x=600,y=(y+3)*(self.height-100)/len(self.isimler)+(self.height-100)/(len(self.isimler)*10),width=75,height=((self.height-100)/(len(self.isimler)*10/8)))
        self.entryler.append(entry)
        self.canvas.create_rectangle(0,y*(self.height-100)/len(self.isimler),600,(y+3)*(self.height-100)/len(self.isimler),fill="#AAAAAA")
        self.canvas.create_text(10,((y+2)*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text=str(self.isimler[y]),anchor=W)
        self.canvas.create_text(550,((y+2)*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text=str(self.azamiler[y]))
        self.canvas.create_text(10,((y+6)*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text="PERSONELE VERİLEN PUAN TOPLAMI",anchor=W)
        self.canvas.create_line(500,0,500,self.height-50)
        self.canvas.create_line(599,0,599,self.height-50)
        self.entryler[8].config(state="readonly")
        self.entryler[19].config(state="readonly")
        self.entryler[22].config(state="readonly")
        self.entryler[27].config(state="readonly")
        self.entryler[33].config(state="readonly")
        self.canvas.create_text(785,20,text="Değerlendirilecek Personel İsmi",fill="white")
        self.seçme = Combobox(self.canvas)
        self.seçme["state"] = "readonly"
        self.seçme.bind("<<ComboboxSelected>>",self.yeni_seçildi)
        self.seçme.place(x=725,y=40,width=125)
        self.kaydet_button = Button(self.canvas,text="Kaydet",command=self.kaydet_basıldı)
        self.kaydet_button.place(x=738,y=90,width=100,height=30)
        self.yazdır_button = Button(self.canvas,text="Yazdır",command=self.yazdır)
        self.yazdır_button.place(x=738,y=250,width=100,height=30)
        self.tarih_entry = Entry(self.canvas)
        self.tarih_entry.place(x=725,y=200,width=125)
        self.canvas.create_text(785,185,text="Tarih",fill="white")
        

    def yazdır(self):
        if self.seçme.current()!=0:
            kişi = self.parent.pencereler["ödül pencere"].kişiler[self.seçme.current()-1]
            wb = pyx.load_workbook("Puan Form.xlsx")
            ws = wb["Form"]
            ws["B4"].value = str(kişi.bilgiler[0])
            ws["B5"].value = str(kişi.bilgiler[1])
            ws["B6"].value = str(kişi.bilgiler[2])
            ws["B7"].value = str(kişi.bilgiler[3])
            ws["B8"].value = str(kişi.bilgiler[4])
            ws["B9"].value = str(kişi.bilgiler[5])
            ws["B10"].value = str(kişi.bilgiler[6])
            for i in range(13,46):
                if i-13<len(kişi.puanlar[0]):
                    ws["C"+str(i)].value = str(kişi.puanlar[0][i-13])
            ws["A50"].value = str(self.tarih_entry.get())
            ws["A51"].value = str(kişi.bilgiler[10])
            ws["A52"].value = str(kişi.bilgiler[11])
            filename = fd.asksaveasfilename(initialfile=f"Disiplin Amiri {kişi.bilgiler[0]} Form.xlsx",defaultextension=".xlsx",filetypes=[("Excel Files","*.xlsx")])
            if filename=="":
                return
            try:
                wb.save(filename)
                os.startfile(filename)
            except:
                messagebox.showerror("Hata","Excel Dosyasını Kapatmadınız")

    def yeni_seçildi(self,event):
        for i in range(len(self.entryler)):
            self.entryler[i].delete(0,END)
            if self.seçme.current()!=0:
                kişi = self.parent.pencereler["ödül pencere"].kişiler[self.seçme.current()-1]    
                if i<len(kişi.puanlar[0]):
                    self.entryler[i].insert(0,kişi.puanlar[0][i])

    def kaydet_basıldı(self):
        if self.seçme.current()!=0:
            kişi = self.parent.pencereler["ödül pencere"].kişiler[self.seçme.current()-1]
            kişi.puanlar[0] = []
            for i in self.entryler:
                kişi.puanlar[0].append(i.get())
            self.parent.pencereler["ödül pencere"].kayıt_yap()
        

    def değişim(self,event):
        toplam = 0
        for i in self.entryler[:-1]:
            if i.get()!="":
                try:
                    toplam += int(i.get())
                except:
                    pass
        self.entryler[-1].config(state="normal")
        self.entryler[-1].delete(0,END)
        self.entryler[-1].insert(0,str(toplam))
        

    def kapat(self):
        self.pencere.withdraw()
        self.parent.pencere.deiconify()
        self.seçme.current(0)
        for i in self.entryler:
            i.delete(0,END)


    def show(self):
        self.canvas.update()
        self.pencere.update()

    def update(self):
        self.seçme["values"] = [""]+[i.bilgiler[0] for i in self.parent.pencereler["ödül pencere"].kişiler]
        self.show()
        
class ilçe_pencere:
    def __init__(self,parent):
        self.parent = parent
        self.pencere = Tk()
        self.pencere.resizable(0,0)
        self.pencere.title("Ödül Değerlendirme Programı")
        self.pencere.geometry("+300+0")
        self.width = 900
        self.height = 650
        self.canvas = Canvas(self.pencere,width=self.width,height=self.height,highlightthickness=0,bg="#424744")
        self.canvas.pack()
        self.pencere.withdraw()
        self.pencere.iconbitmap('Bayrak.ico')
        self.pencere.protocol("WM_DELETE_WINDOW", self.kapat)
        self.isimler = ["Genel Nitelikler","Çevresi ile ilişkileri","Görevin gerektirdiği sır saklama ve gizliliğe riayeti",
                        "Görevinde tarafsızlık ve eşitlik ilkesine bağlılığı","Görevini benimseme ve sorumluluk alma özelliği","Genel kültür bilgisi",
                        "Görevini yapmadaki hevesi ve kendini yetiştirme gayreti","Bireysel iş görme kabiliyeti ve teşebbüs fikri",
                        "Meslek şeref ve haysiyetini zedeleyici kötü alışkanlıklara düşkün olmaması","Mesleki Yeterlilik","Mesleki bilgisi ve uygulamaya yansıtması",
                        "Görevini yapmadaki başarısı","Göreve başlamadan önce hazırlık yapma becerisi","Görevinin gerektirdiği faaliyetleri gerçekleştirmesi",
                        "Kullandığı eşyanın bakım, koruma, tertip ve düzenini sağlama","İşbirliği yapmada ve değişen şartlara, görevlere uymada gösterdiği başarı",
                        "Mevzuat Bilgisi, bağlılığı ve uygulaması","Mesleki, sosyal ve kültürel faliyet durumu",
                        "Görevi sırasında ortaya çıkan problemleri pratik ve akılcı bir şekilde çözümlemedeki başarısı",
                        "Uzun süreli çalışmalarda fikri ve bedeni dayanıklılığı","Mesleği ile İlgili Faaliyetleri, Olağanüstü Gayret ve Çalışmaları ile Diğer Hususlar",
                        "Gazete ve dergilerde yayımlanmış bilimsel ve mesleki inceleme,araştırma ve makaleler İçin","Yazmış olduğu (ISBN almış) kitapları için",
                        "Türkiye genelinde yapılan yarışmalarda;","a. Birinci olanlara","b. İkinci olanlara","c. Üçüncü olanlara","d. Mansiyon alanlara",
                        "Yurt dışı yarışmaların her birinde; ","a. Birinci olanlara","b. İkinci olanlara","c. Üçüncü olanlara","d. Mansiyon alanlara",
                        """Mesleği ile ilgili bir buluş yaparak veya mevcut usullerde yararlı
yenilikler meydana getirerek veyahut başka yolarla gelir ve verimliliğin arttırılmasında yahut
hizmetin yürütülmesinde dikkate değer çalışmalar yapmak (tescil edilmiş olmalı)"""]
        self.azamiler = ["Azami Puan",5,5,5,5,5,5,5,5,"",5,5,5,5,5,5,5,5,5,5,"",6,8,"",8,6,4,2,"",10,8,6,4,10,""]
        self.entryler = []
        self.engel = [1,2,3,4,9,10,11,12,13,14,15,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34]
        self.canvas.create_rectangle(0,0,675,self.height-50,fill="#AAAAAA")
        self.canvas.create_text(624,7,text="Puan",anchor=W)
        for y in range(0,len(self.isimler)):
            self.canvas.create_rectangle(0,y*(self.height-100)/len(self.isimler),599,(y+1)*(self.height-100)/len(self.isimler),fill="#AAAAAA")
            if y!=33:
                self.canvas.create_rectangle(599,y*(self.height-100)/len(self.isimler),675,(y+1)*(self.height-100)/len(self.isimler),fill="#AAAAAA")
            if y!=len(self.isimler)-1:
                self.canvas.create_text(10,(y*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text=str(self.isimler[y]),#+f" (Azami Puan:{str(self.azamiler[y])})"
                                anchor=W)
                self.canvas.create_text(550,(y*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text=str(self.azamiler[y]),#+f" (Azami Puan:{str(self.azamiler[y])})"
                )
                if y!=0:
                    entry = Entry(self.canvas)
                    entry.bind("<KeyRelease>", self.değişim)  
                    if not y in self.engel:
                        entry.place(x=600,y=y*(self.height-100)/len(self.isimler)+(self.height-100)/(len(self.isimler)*10),width=75,height=(self.height-100)/(len(self.isimler)*10/8))
                    self.entryler.append(entry)
        entry = Entry(self.canvas)
        entry.bind("<KeyRelease>", self.değişim)
        #entry.place(x=600,y=(y)*(self.height-100)/len(self.isimler)+(self.height-100)/(len(self.isimler)*10),width=75,height=((self.height-100)/(len(self.isimler)*10/8))*3.5)
        self.entryler.append(entry)
        entry = Entry(self.canvas)
        entry.bind("<KeyRelease>", self.değişim)
        entry.place(x=600,y=(y+3)*(self.height-100)/len(self.isimler)+(self.height-100)/(len(self.isimler)*10),width=75,height=((self.height-100)/(len(self.isimler)*10/8)))
        self.entryler.append(entry)
        self.canvas.create_rectangle(0,y*(self.height-100)/len(self.isimler),600,(y+3)*(self.height-100)/len(self.isimler),fill="#AAAAAA")
        self.canvas.create_rectangle(599,y*(self.height-100)/len(self.isimler),675,(y+3)*(self.height-100)/len(self.isimler),fill="#AAAAAA")
        self.canvas.create_text(10,((y+2)*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text=str(self.isimler[y]),anchor=W)
        self.canvas.create_text(550,((y+2)*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text=str(self.azamiler[y]))
        self.canvas.create_text(10,((y+6)*(self.height-100)/len(self.isimler)+(y+1)*(self.height-100)/len(self.isimler))/2,text="PERSONELE VERİLEN PUAN TOPLAMI",anchor=W)
        self.canvas.create_line(500,0,500,self.height-50)
        self.canvas.create_line(599,0,599,self.height-50)
        self.entryler[0].config(state="readonly")
        self.entryler[1].config(state="readonly")
        self.entryler[2].config(state="readonly")
        self.entryler[3].config(state="readonly")
        self.entryler[8].config(state="readonly")
        self.entryler[9].config(state="readonly")
        self.entryler[10].config(state="readonly")
        self.entryler[11].config(state="readonly")
        self.entryler[12].config(state="readonly")
        self.entryler[13].config(state="readonly")
        self.entryler[14].config(state="readonly")
        self.entryler[19].config(state="readonly")
        for i in range(20,len(self.entryler)):
            self.entryler[i].config(state="readonly")
        self.entryler[22].config(state="readonly")
        self.entryler[27].config(state="readonly")
        self.entryler[33].config(state="readonly")
        self.canvas.create_text(785,20,text="Değerlendirilecek Personel İsmi",fill="white")
        self.seçme = Combobox(self.canvas)
        self.seçme["state"] = "readonly"
        self.seçme.bind("<<ComboboxSelected>>",self.yeni_seçildi)
        self.seçme.place(x=725,y=40,width=125)
        self.kaydet_button = Button(self.canvas,text="Değerlendirmeyi\nKaydet",command=self.kaydet_basıldı)
        self.kaydet_button.place(x=738,y=90,width=100,height=30)
        self.canvas.create_text(785,150,text="Komisyon Üyesi İsmi",fill="white")
        self.komisyon_seçme = Combobox(self.canvas)
        self.komisyon_seçme["state"] = "readonly"
        self.komisyon_seçme.bind("<<ComboboxSelected>>",self.yeni_seçildi)
        self.komisyon_seçme.place(x=725,y=170,width=125)
        self.kaydet_button = Button(self.canvas,text="Komisyon Üyesi\n Ekle",command=self.üye_ekle)
        self.kaydet_button.place(x=738,y=220,width=100,height=30)
        self.kaydet_button = Button(self.canvas,text="Komisyon Üyes\n Sil",command=self.üye_sil)
        self.kaydet_button.place(x=738,y=295,width=100,height=30)
        self.yazdır_button = Button(self.canvas,text="Seçili Kom. Üyesi\n Bireysel Yazdır",command=self.yazdır)
        self.yazdır_button.place(x=738,y=430,width=100,height=40)
        self.yazdır_button = Button(self.canvas,text="Toplu Yazdır",command=lambda :self.ortalama_yazdır(True),bg="Gray")
        self.yazdır_button.place(x=738,y=570,width=100,height=40)
        self.ilçe_yazdır_button = Button(self.canvas,text="Seçili Personel\nPuan Ort. Yazdır",command=self.ortalama_yazdır)
        self.ilçe_yazdır_button.place(x=738,y=500,width=100,height=45)
        self.tarih_entry = Entry(self.canvas)
        self.tarih_entry.place(x=725,y=390)
        self.canvas.create_text(785,370,text="Tarih",fill="white")

    def yazdır(self):
        if self.seçme.current()!=0 and self.komisyon_seçme.current()!=0:
            kişi = self.parent.pencereler["ödül pencere"].kişiler[self.seçme.current()-1]
            wb = pyx.load_workbook("Puan Form.xlsx")
            ws = wb["Form"]
            ws["B4"].value = str(kişi.bilgiler[0])
            ws["B5"].value = str(kişi.bilgiler[1])
            ws["B6"].value = str(kişi.bilgiler[2])
            ws["B7"].value = str(kişi.bilgiler[3])
            ws["B8"].value = str(kişi.bilgiler[4])
            ws["B9"].value = str(kişi.bilgiler[5])
            ws["B10"].value = str(kişi.bilgiler[6])
            for i in range(13,46):
                if i-13<len(kişi.puanlar[0]):
                    ws["C"+str(i)].value = str(kişi.puanlar[0][i-13])
            ws["A50"].value = str(self.tarih_entry.get())
            ws["A51"].value = str(kişi.bilgiler[10])
            ws["A52"].value = str(kişi.bilgiler[11])
            for i in range(13,46):
                if str(kişi.puanlar[1][self.komisyon_seçme.current()-1][i-13])!="":
                    ws["D"+str(i)].value = str(kişi.puanlar[1][self.komisyon_seçme.current()-1][i-13])
            ws["B50"].value = str(self.tarih_entry.get())
            ws["B51"].value = str(self.komisyon_seçme.get())
            ws["B52"].value = str(self.parent.pencereler["ödül pencere"].unvanlar[self.komisyon_seçme.current()-1])
            filename = fd.asksaveasfilename(initialfile=f"İlçe {kişi.bilgiler[0]} {str(self.komisyon_seçme.get())} Form.xlsx",defaultextension=".xlsx",filetypes=[("Excel Files","*.xlsx")])
            if filename=="":
                return
            try:
                wb.save(filename)
                os.startfile(filename)
            except:
                messagebox.showerror("Hata","Excel Dosyasını Kapatmadınız")

    def ortalama_yazdır(self,toplu=False):
        if toplu:
            wb = pyx.load_workbook("Puan Form.xlsx")
            for kişi in self.parent.pencereler["ödül pencere"].kişiler:
                ws = wb.copy_worksheet(wb["Form"])
                ws.title = str(kişi.bilgiler[0])
                ws["B4"].value = str(kişi.bilgiler[0])
                ws["B5"].value = str(kişi.bilgiler[1])
                ws["B6"].value = str(kişi.bilgiler[2])
                ws["B7"].value = str(kişi.bilgiler[3])
                ws["B8"].value = str(kişi.bilgiler[4])
                ws["B9"].value = str(kişi.bilgiler[5])
                ws["B10"].value = str(kişi.bilgiler[6])
                for i in range(13,46):
                    if i-13<len(kişi.puanlar[0]):
                        ws["C"+str(i)].value = str(kişi.puanlar[0][i-13])
                ws["A50"].value = str(self.tarih_entry.get())
                ws["A51"].value = str(kişi.bilgiler[10])
                ws["A52"].value = str(kişi.bilgiler[11])
                üyeler = []
                unvanları = []
                indexler = []
                for index,i in enumerate(kişi.puanlar[1]):
                    boş = True
                    for y in i:
                        if y!="":
                            try:
                                d = int(y)
                            except:
                                continue
                            if d>0:
                                boş = False
                                break
                    if not boş:
                        üyeler.append(self.parent.pencereler["ödül pencere"].komisyon[index])
                        unvanları.append(self.parent.pencereler["ödül pencere"].unvanlar[index])
                        indexler.append(index)
                hepsi = 0
                for i in range(13,46):
                    toplam = 0
                    sayı = 0
                    for index in indexler:
                        if str(kişi.puanlar[1][index][i-13])!="":
                            try:
                                val = int(kişi.puanlar[1][index][i-13])
                            except:
                                continue
                            sayı += 1
                            toplam += val
                    if sayı!=0:
                        toplam = toplam/sayı
                    if toplam!=0:
                        ws["D"+str(i)].value = toplam
                    sayı = 0
                    val2 = 0
                    if toplam!=0:
                        sayı += 1
                    if i-13<len(kişi.puanlar[0]) and kişi.puanlar[0][i-13]!="":
                        try:
                            d = int(kişi.puanlar[0][i-13])
                            val2 = d
                        except:
                            pass
                        if val2!=0:
                            sayı+=1
                    if sayı!=0:
                        ws["G"+str(i)].value = (toplam+val2)/sayı
                        hepsi += (toplam+val2)/sayı
                ws["G46"].value = hepsi
                ws["C57"].value = hepsi
                ws["B50"].value = str(self.tarih_entry.get())
                if len(üyeler)>0:
                    ws["B51"].value = üyeler[0]
                    ws["B52"].value = unvanları[0]
                if len(üyeler)>1:
                    ws["D51"].value = üyeler[1]
                    ws["D52"].value = unvanları[1]
                if len(üyeler)>2:
                    ws["F51"].value = üyeler[2]
                    ws["F52"].value = unvanları[2]
                if len(üyeler)>3:
                    ws["C54"].value = üyeler[3]
                    ws["C55"].value = unvanları[3]
                if len(üyeler)>4:
                    ws["E54"].value = üyeler[4]
                    ws["E55"].value = unvanları[4]
            wb.remove(wb["Form"])
            filename = fd.asksaveasfilename(initialfile=f"Toplu İl-İlçe Form.xlsx",defaultextension=".xlsx",filetypes=[("Excel Files","*.xlsx")])
            if filename=="":
                return
            try:
                wb.save(filename)
                os.startfile(filename)
            except:
                messagebox.showerror("Hata","Excel Dosyasını Kapatmadınız")
            return 
        if self.seçme.current()!=0:
            kişi = self.parent.pencereler["ödül pencere"].kişiler[self.seçme.current()-1]
            wb = pyx.load_workbook("Puan Form.xlsx")
            ws = wb["Form"]
            ws["B4"].value = str(kişi.bilgiler[0])
            ws["B5"].value = str(kişi.bilgiler[1])
            ws["B6"].value = str(kişi.bilgiler[2])
            ws["B7"].value = str(kişi.bilgiler[3])
            ws["B8"].value = str(kişi.bilgiler[4])
            ws["B9"].value = str(kişi.bilgiler[5])
            ws["B10"].value = str(kişi.bilgiler[6])
            for i in range(13,46):
                if i-13<len(kişi.puanlar[0]):
                    ws["C"+str(i)].value = str(kişi.puanlar[0][i-13])
            ws["A50"].value = str(self.tarih_entry.get())
            ws["A51"].value = str(kişi.bilgiler[10])
            ws["A52"].value = str(kişi.bilgiler[11])
            üyeler = []
            unvanları = []
            indexler = []
            for index,i in enumerate(kişi.puanlar[1]):
                boş = True
                for y in i:
                    if y!="":
                        try:
                            d = int(y)
                        except:
                            continue
                        if d>0:
                            boş = False
                            break
                if not boş:
                    üyeler.append(self.parent.pencereler["ödül pencere"].komisyon[index])
                    unvanları.append(self.parent.pencereler["ödül pencere"].unvanlar[index])
                    indexler.append(index)
            hepsi = 0
            for i in range(13,46):
                toplam = 0
                sayı = 0
                for index in indexler:
                    if str(kişi.puanlar[1][index][i-13])!="":
                        try:
                            val = int(kişi.puanlar[1][index][i-13])
                        except:
                            continue
                        sayı += 1
                        toplam += val
                if sayı!=0:
                    toplam = toplam/sayı
                if toplam!=0:
                    ws["D"+str(i)].value = toplam
                sayı = 0
                val2 = 0
                if toplam!=0:
                    sayı += 1
                if i-13<len(kişi.puanlar[0]) and kişi.puanlar[0][i-13]!="":
                    try:
                        d = int(kişi.puanlar[0][i-13])
                        val2 = d
                    except:
                        pass
                    if val2!=0:
                        sayı+=1
                if sayı!=0:
                    ws["G"+str(i)].value = (toplam+val2)/sayı
                    hepsi += (toplam+val2)/sayı
            ws["G46"].value = hepsi
            ws["C57"].value = hepsi
            ws["B50"].value = str(self.tarih_entry.get())
            if len(üyeler)>0:
                ws["B51"].value = üyeler[0]
                ws["B52"].value = unvanları[0]
            if len(üyeler)>1:
                ws["D51"].value = üyeler[1]
                ws["D52"].value = unvanları[1]
            if len(üyeler)>2:
                ws["F51"].value = üyeler[2]
                ws["F52"].value = unvanları[2]
            if len(üyeler)>3:
                ws["C54"].value = üyeler[3]
                ws["C55"].value = unvanları[3]
            if len(üyeler)>4:
                ws["E54"].value = üyeler[4]
                ws["E55"].value = unvanları[4]
            filename = fd.asksaveasfilename(initialfile=f"İlçe {kişi.bilgiler[0]} Form.xlsx",defaultextension=".xlsx",filetypes=[("Excel Files","*.xlsx")])
            if filename=="":
                return
            try:
                wb.save(filename)
                os.startfile(filename)
            except:
                messagebox.showerror("Hata","Excel Dosyasını Kapatmadınız")

    def üye_sil(self):
        if self.komisyon_seçme.current()!=0:
            answer = messagebox.askquestion("Eminmisiniz","Eminmisiniz")
            if answer=="yes":
                self.parent.pencereler["ödül pencere"].komisyon.pop(self.komisyon_seçme.current()-1)
                self.parent.pencereler["ödül pencere"].unvanlar.pop(self.komisyon_seçme.current()-1)
                for i in self.parent.pencereler["ödül pencere"].kişiler:
                    i.puanlar[1].pop(self.komisyon_seçme.current()-1)
                self.komisyon_seçme.current(0)
                for i in self.entryler:
                    i.delete(0,END)
                self.parent.pencereler["ödül pencere"].kayıt_yap()
                self.update()

    def üye_ekle(self):
        isim = tkinter.simpledialog.askstring("Komisyon Üyesi İsim Soyisim","Komisyon Üyesi İsim Soyisim")
        if isim!=None:
            unvan = tkinter.simpledialog.askstring("Komisyon Üyesi Unvanı","Komisyon Üyesi Unvanı")
            if unvan!=None:
                self.parent.pencereler["ödül pencere"].komisyon.append(isim)
                self.parent.pencereler["ödül pencere"].unvanlar.append(unvan)
                for i in self.parent.pencereler["ödül pencere"].kişiler:
                    i.puanlar[1].append([])
                self.parent.pencereler["ödül pencere"].kayıt_yap()
                self.update()
            

    def yeni_seçildi(self,event):
        for i in range(len(self.entryler)):
            self.entryler[i].delete(0,END)
            if self.seçme.current()!=0 and self.komisyon_seçme.current()!=0:
                kişi = self.parent.pencereler["ödül pencere"].kişiler[self.seçme.current()-1]
                if i<len(kişi.puanlar[1][self.komisyon_seçme.current()-1]):
                    self.entryler[i].insert(0,kişi.puanlar[1][self.komisyon_seçme.current()-1][i])

    def kaydet_basıldı(self):
        if self.seçme.current()!=0 and self.komisyon_seçme.current()!=0:
            kişi = self.parent.pencereler["ödül pencere"].kişiler[self.seçme.current()-1]
            kişi.puanlar[1][self.komisyon_seçme.current()-1] = []
            for i in self.entryler:
                kişi.puanlar[1][self.komisyon_seçme.current()-1].append(i.get())
            self.parent.pencereler["ödül pencere"].kayıt_yap()
        

    def değişim(self,event):
        toplam = 0
        for i in self.entryler[:-1]:
            if i.get()!="":
                try:
                    toplam += int(i.get())
                except:
                    pass
        self.entryler[-1].config(state="normal")
        self.entryler[-1].delete(0,END)
        self.entryler[-1].insert(0,str(toplam))
        

    def kapat(self):
        self.pencere.withdraw()
        self.parent.pencere.deiconify()
        self.seçme.current(0)
        for i in self.entryler:
            i.delete(0,END)

    def show(self):
        self.canvas.update()
        self.pencere.update()

    def update(self):
        self.seçme["values"] = [""]+[i.bilgiler[0] for i in self.parent.pencereler["ödül pencere"].kişiler]
        self.komisyon_seçme["values"] = [""]+self.parent.pencereler["ödül pencere"].komisyon
        self.show()

class hakkında_pencere:
    def __init__(self,parent):
        self.eski_seçili_index = 0
        self.kişiler = []
        self.komisyon = []
        self.unvanlar = []
        self.parent = parent
        self.pencere = Tk()
        self.pencere.resizable(0,0)
        self.pencere.geometry("+300+0")
        self.pencere.title("Ödül Değerlendirme Programı")
        self.width = 300
        self.height = 160
        self.canvas = Canvas(self.pencere,width=self.width,height=self.height,highlightthickness=0,bg="White")
        self.canvas.pack()
        self.pencere.withdraw()
        self.pencere.iconbitmap('Bayrak.ico')
        self.pencere.protocol("WM_DELETE_WINDOW", self.kapat)
        self.canvas.create_text(150,60,text="Hazırlayanlar",font=("italic",10,"bold"))
        self.canvas.create_text(150,90,text="Rasül ÇÖVÜT",font=("italic",10,"bold"))
        self.canvas.create_text(150,110,text="Mehmet KARAŞAH",font=("italic",10,"bold"))
        self.canvas.create_text(150,130,text="Ömer Selim KARAŞAH",font=("italic",10,"bold"))
        self.canvas.create_text(280,150,text="2022",font=("italic",10,"bold"),anchor="e")
        self.canvas.create_text(150,10,text="Denizli Bozkurt İlçe Milli Eğitim Müdürlüğü",font=("italic",10,"bold"))
        self.canvas.create_text(150,30,text="Ödül Değerlendirme Programı",font=("italic",10,"bold"))# 


    def kapat(self):
        self.pencere.withdraw()
        self.parent.pencereler["ödül pencere"].pencere.deiconify()

    def show(self):
        self.canvas.update()
        self.pencere.update()

    def update(self):
        self.show()

an = time.time()
ana = ana_pencere()
ana.show()
ana.pencere.mainloop()
    
